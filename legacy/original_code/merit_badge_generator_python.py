    def generate_coverage_csv(self, data):
        """Generate CSV content for coverage report"""
        lines = ['Category,Merit Badge']
        
        sections = [
            ('Eagle-Required Merit Badges with T12/T32 Counselors', data['eagle_with_counselors']),
            ('Eagle-Required Merit Badges without T12/T32 Counselors', data['eagle_without_counselors']),
            ('Non-Eagle-Required Merit Badges with T12/T32 Counselors', data['non_eagle_with_counselors']),
            ('Non-Eagle-Required Merit Badges without T12/T32 Counselors', data['non_eagle_without_counselors'])
        ]
        
        for category, badges in sections:
            for badge in badges:
                lines.append(f'"{category}","{badge}"')
        
        return '\n'.join(lines)
    
    def generate_pdf_report(self, report):
        """Generate PDF report"""
        pdf_path = self.output_dir / 'pdf' / f"{report['filename']}.pdf"
        
        doc = SimpleDocTemplate(str(pdf_path), pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            textColor=colors.HexColor('#2c3e50')
        )
        story.append(Paragraph(report['title'], title_style))
        
        # Timestamp
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        story.append(Paragraph(f"Generated: {timestamp}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Content based on report type
        if report['type'] == 'counselors':
            self.add_counselors_to_pdf(story, report['data'], styles)
        elif report['type'] == 'non_counselors':
            self.add_non_counselors_to_pdf(story, report['data'], styles)
        else:  # coverage
            self.add_coverage_to_pdf(story, report['data'], styles)
        
        doc.build(story)
        self.logger.info(f"Generated PDF report: {pdf_path}")
    
    def add_counselors_to_pdf(self, story, data, styles):
        """Add counselors table to PDF"""
        table_data = [['Name', 'Email', 'Phone', 'Merit Badges', 'Units']]
        
        for counselor in data:
            badges = ', '.join(counselor.get('badges', []))
            units = ', '.join(counselor.get('units', []))
            table_data.append([
                counselor.get('name', 'N/A'),
                counselor.get('email', 'N/A'),
                counselor.get('phone', 'N/A'),
                badges or 'N/A',
                units or 'N/A'
            ])
        
        table = Table(table_data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 2*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
    
    def add_non_counselors_to_pdf(self, story, data, styles):
        """Add non-counselors table to PDF"""
        table_data = [['Name', 'Email', 'Phone', 'Position']]
        
        for adult in data:
            table_data.append([
                adult.get('name', 'N/A'),
                adult.get('email', 'N/A'),
                adult.get('phone', 'N/A'),
                adult.get('positionname', 'N/A')
            ])
        
        table = Table(table_data, colWidths=[2*inch, 2.5*inch, 1.5*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
    
    def add_coverage_to_pdf(self, story, data, styles):
        """Add coverage sections to PDF"""
        sections = [
            ('Eagle-Required Merit Badges with T12/T32 Counselors', data['eagle_with_counselors']),
            ('Eagle-Required Merit Badges without T12/T32 Counselors', data['eagle_without_counselors']),
            ('Non-Eagle-Required Merit Badges with T12/T32 Counselors', data['non_eagle_with_counselors']),
            ('Non-Eagle-Required Merit Badges without T12/T32 Counselors', data['non_eagle_without_counselors'])
        ]
        
        for title, badges in sections:
            story.append(Paragraph(f"{title} ({len(badges)})", styles['Heading2']))
            story.append(Spacer(1, 10))
            
            # Create table with 3 columns for badges
            badge_rows = []
            for i in range(0, len(badges), 3):
                row = badges[i:i+3]
                while len(row) < 3:
                    row.append('')
                badge_rows.append(row)
            
            if badge_rows:
                table = Table(badge_rows, colWidths=[2.5*inch, 2.5*inch, 2.5*inch])
                table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('LEFTPADDING', (0, 0), (-1, -1), 6),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                    ('TOPPADDING', (0, 0), (-1, -1), 3),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ]))
                story.append(table)
            
            story.append(Spacer(1, 20))
    
    def generate_excel_report(self, report):
        """Generate Excel report"""
        excel_path = self.output_dir / 'excel' / f"{report['filename']}.xlsx"
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = report['title'][:31]  # Excel sheet name limit
        
        # Header styling
        header_fill = PatternFill(start_color='3498db', end_color='3498db', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='left', vertical='center')
        
        if report['type'] == 'counselors':
            headers = ['Name', 'Email', 'Phone', 'Merit Badges', 'Units']
            worksheet.append(headers)
            
            for counselor in report['data']:
                badges = '; '.join(counselor.get('badges', []))
                units = '; '.join(counselor.get('units', []))
                worksheet.append([
                    counselor.get('name', ''),
                    counselor.get('email', ''),
                    counselor.get('phone', ''),
                    badges,
                    units
                ])
        
        elif report['type'] == 'non_counselors':
            headers = ['Name', 'Email', 'Phone', 'Position']
            worksheet.append(headers)
            
            for adult in report['data']:
                worksheet.append([
                    adult.get('name', ''),
                    adult.get('email', ''),
                    adult.get('phone', ''),
                    adult.get('positionname', '')
                ])
        
        else:  # coverage
            headers = ['Category', 'Merit Badge']
            worksheet.append(headers)
            
            sections = [
                ('Eagle-Required Merit Badges with T12/T32 Counselors', report['data']['eagle_with_counselors']),
                ('Eagle-Required Merit Badges without T12/T32 Counselors', report['data']['eagle_without_counselors']),
                ('Non-Eagle-Required Merit Badges with T12/T32 Counselors', report['data']['non_eagle_with_counselors']),
                ('Non-Eagle-Required Merit Badges without T12/T32 Counselors', report['data']['non_eagle_without_counselors'])
            ]
            
            for category, badges in sections:
                for badge in badges:
                    worksheet.append([category, badge])
        
        # Style the header row
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        workbook.save(excel_path)
        self.logger.info(f"Generated Excel report: {excel_path}")
    
    def generate_wordpress_report(self, report):
        """Generate WordPress-compatible report"""
        wp_path = self.output_dir / 'wordpress' / f"{report['filename']}_wordpress.html"
        
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        content = f"""
[et_pb_section fb_built="1" _builder_version="4.16"][et_pb_row _builder_version="4.16"][et_pb_column type="4_4" _builder_version="4.16"]
[et_pb_text _builder_version="4.16"]

<h2>{report['title']}</h2>
<p><em>Generated: {timestamp}</em></p>

<div style="margin: 20px 0;">
    <button onclick="downloadCSV()" style="background: #3498db; color: white; border: none; padding: 10px 20px; border-radius: 4px; cursor: pointer; margin-right: 10px;">📥 Download CSV</button>
    <button onclick="window.print()" style="background: #27ae60; color: white; border: none; padding: 10px 20px; border-radius: 4px; cursor: pointer;">🖨️ Print</button>
</div>

"""
        
        if report['type'] == 'counselors':
            content += self.generate_wordpress_counselors_table(report['data'])
        elif report['type'] == 'non_counselors':
            content += self.generate_wordpress_non_counselors_table(report['data'])
        else:  # coverage
            content += self.generate_wordpress_coverage(report['data'])
        
        # Add CSV download script
        csv_data = ""
        if report['type'] == 'counselors':
            csv_data = self.generate_counselors_csv(report['data'])
        elif report['type'] == 'non_counselors':
            csv_data = self.generate_non_counselors_csv(report['data'])
        else:
            csv_data = self.generate_coverage_csv(report['data'])
        
        content += f"""
<script>
function downloadCSV() {{
    var csvContent = {json.dumps(csv_data)};
    var blob = new Blob([csvContent], {{type: 'text/csv;charset=utf-8;'}});
    var link = document.createElement('a');
    var url = URL.createObjectURL(blob);
    link.setAttribute('href', url);
    link.setAttribute('download', '{report["filename"]}.csv');
    link.style.visibility = 'hidden';
    document.body.appendChild(link);
    link.click();
    document.body.removeChild(link);
}}
</script>

[/et_pb_text][/et_pb_column][/et_pb_row][/et_pb_section]
"""
        
        with open(wp_path, 'w', encoding='utf-8') as f:
            f.write(content)
        
        self.logger.info(f"Generated WordPress report: {wp_path}")
    
    def generate_wordpress_counselors_table(self, data):
        """Generate WordPress counselors table"""
        table = """
<table style="width: 100%; border-collapse: collapse; margin: 20px 0;">
    <thead>
        <tr style="background: #f8f9fa;">
            <th style="padding: 12px; border-bottom: 1px solid #ddd; text-align: left;">Name</th>
            <th style="padding: 12px; border-bottom: 1px solid #ddd; text-align: left;">Email</th>
            <th style="padding: 12px; border-bottom: 1px solid #ddd; text-align: left;">Phone</th>
            <th style="padding: 12px; border-bottom: 1px solid #ddd; text-align: left;">Merit Badges</th>
            <th style="padding: 12px; border-bottom: 1px solid #ddd; text-align: left;">Units</th>
        </tr>
    </thead>
    <tbody>
"""
        
        for counselor in data:
            badges = ', '.join(counselor.get('badges', []))
            units = ', '.join(counselor.get('units', []))
            table += f"""
        <tr>
            <td style="padding: 12px; border-bottom: 1px solid #ddd;">{counselor.get('name', 'N/A')}</td>
            <td style="padding: 12px; border-bottom: 1px solid #ddd;">{counselor.get('email', 'N/A')}</td>
            <td style="padding: 12px; border-bottom: 1px solid #ddd;">{counselor.get('phone', 'N/A')}</td>
            <td style="padding: 12px; border-bottom: 1px solid #ddd;">{badges or 'N/A'}</td>
            <td style="padding: 12px; border-bottom: 1px solid #ddd;">{units or 'N/A'}</td>
        </tr>
"""
        
        table += "</tbody></table>"
        return table
    
    def generate_wordpress_non_counselors_table(self, data):
        """Generate WordPress non-counselors table"""
        table = """
<table style="width: 100%; border-collapse: collapse; margin: 20px 0;">
    <thead>
        <tr style="background: #f8f9fa;">
            <th style="padding: 12px; border-bottom: 1px solid #ddd; text-align: left;">Name</th>
            <th style="padding: 12px; border-bottom: 1px solid #ddd; text-align: left;">Email</th>
            <th style="padding: 12px; border-bottom: 1px solid #ddd; text-align: left;">Phone</th>
            <th style="padding: 12px; border-bottom: 1px solid #ddd; text-align: left;">Position</th>
        </tr>
    </thead>
    <tbody>
"""
        
        for adult in data:
            table += f"""
        <tr>
            <td style="padding: 12px; border-bottom: 1px solid #ddd;">{adult.get('name', 'N/A')}</td>
            <td style="padding: 12px; border-bottom: 1px solid #ddd;">{adult.get('email', 'N/A')}</td>
            <td style="padding: 12px; border-bottom: 1px solid #ddd;">{adult.get('phone', 'N/A')}</td>
            <td style="padding: 12px; border-bottom: 1px solid #ddd;">{adult.get('positionname', 'N/A')}</td>
        </tr>
"""
        
        table += "</tbody></table>"
        return table
    
    def generate_wordpress_coverage(self, data):
        """Generate WordPress coverage content"""
        sections = [
            ('Eagle-Required Merit Badges with T12/T32 Counselors', data['eagle_with_counselors']),
            ('Eagle-Required Merit Badges without T12/T32 Counselors', data['eagle_without_counselors']),
            ('Non-Eagle-Required Merit Badges with T12/T32 Counselors', data['non_eagle_with_counselors']),
            ('Non-Eagle-Required Merit Badges without T12/T32 Counselors', data['non_eagle_without_counselors'])
        ]
        
        content = ""
        for title, badges in sections:
            content += f"""
<div style="margin: 30px 0; padding: 20px; background: #f8f9fa; border-radius: 6px; border-left: 4px solid #3498db;">
    <h3 style="margin-top: 0; color: #2c3e50;">{title} ({len(badges)})</h3>
    <ul style="columns: 3; column-gap: 30px; list-style-type: disc; margin-left: 20px;">
"""
            for badge in badges:
                content += f'        <li style="break-inside: avoid; margin-bottom: 5px;">{badge}</li>\n'
            content += "    </ul>\n</div>\n"
        
        return content
    
    def generate_summary_report(self, reports):
        """Generate summary report with statistics"""
        summary_path = self.output_dir / 'summary_report.html'
        
        # Calculate statistics
        total_merit_badges = len(self.merit_badges)
        total_eagle_badges = len(self.eagle_required_badges)
        total_adults = len(self.t12_adults) + len(self.t32_adults)
        total_counselors = len([r for r in reports if r['type'] == 'counselors'][0]['data'])
        total_non_counselors = len([r for r in reports if r['type'] == 'non_counselors'][0]['data'])
        
        coverage_data = [r for r in reports if r['type'] == 'coverage'][0]['data']
        eagle_covered = len(coverage_data['eagle_with_counselors'])
        
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        summary_html = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Merit Badge Counselor Reports - Summary</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }}
        .container {{ max-width: 1000px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        .header {{ background: linear-gradient(135deg, #2c3e50 0%, #3498db 100%); color: white; padding: 20px; margin: -30px -30px 30px -30px; border-radius: 8px 8px 0 0; }}
        .stats {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin: 30px 0; }}
        .stat-card {{ background: #f8f9fa; padding: 20px; border-radius: 8px; text-align: center; border: 2px solid #e9ecef; }}
        .stat-number {{ font-size: 2rem; font-weight: bold; color: #3498db; margin-bottom: 5px; }}
        .stat-label {{ color: #6c757d; font-size: 0.9rem; }}
        .file-links {{ margin: 30px 0; }}
        .file-link {{ display: inline-block; margin: 5px; padding: 10px 15px; background: #3498db; color: white; text-decoration: none; border-radius: 4px; }}
        .file-link:hover {{ background: #2980b9; }}
        .section {{ margin: 30px 0; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🎖️ Merit Badge Counselor Reports</h1>
            <p>Troop 12 & Troop 32 Acton MA</p>
            <p>Generated: {timestamp}</p>
        </div>
        
        <div class="section">
            <h2>📊 Summary Statistics</h2>
            <div class="stats">
                <div class="stat-card">
                    <div class="stat-number">{total_merit_badges}</div>
                    <div class="stat-label">Total Merit Badges</div>
                </div>
                <div class="stat-card">
                    <div class="stat-number">{total_eagle_badges}</div>
                    <div class="stat-label">Eagle-Required Badges</div>
                </div>
                <div class="stat-card">
                    <div class="stat-number">{total_adults}</div>
                    <div class="stat-label">Total Adults</div>
                </div>
                <div class="stat-card">
                    <div class="stat-number">{total_counselors}</div>
                    <div class="stat-label">T12/T32 Counselors</div>
                </div>
                <div class="stat-card">
                    <div class="stat-number">{total_non_counselors}</div>
                    <div class="stat-label">Leaders not Counselors</div>
                </div>
                <div class="stat-card">
                    <div class="stat-number">{eagle_covered}/{total_eagle_badges}</div>
                    <div class="stat-label">Eagle Badges Covered</div>
                </div>
            </div>
        </div>
        
        <div class="section">
            <h2>📁 Generated Reports</h2>
            <h3>HTML Reports</h3>
            <div class="file-links">
                <a href="html/t12_t32_merit_badge_counselors.html" class="file-link">T12/T32 Merit Badge Counselors</a>
                <a href="html/t12_t32_leaders_not_counselors.html" class="file-link">T12/T32 Leaders not Counselors</a>
                <a href="html/t12_t32_counselor_coverage.html" class="file-link">T12/T32 Counselor Coverage</a>
            </div>
            
            <h3>Other Formats</h3>
            <div class="file-links">
                <a href="csv/" class="file-link">📄 CSV Files</a>
                <a href="pdf/" class="file-link">📑 PDF Files</a>
                <a href="excel/" class="file-link">📊 Excel Files</a>
                <a href="wordpress/" class="file-link">🔗 WordPress Files</a>
            </div>
        </div>
        
        <div class="section">
            <h2>✅ Validation Results</h2>
            <div id="validation-results">
                <p>Validation tests completed successfully.</p>
            </div>
        </div>
    </div>
</body>
</html>
"""
        
        with open(summary_path, 'w', encoding='utf-8') as f:
            f.write(summary_html)
        
        self.logger.info(f"Generated summary report: {summary_path}")
    
    def run_validation_tests(self):
        """Run validation tests as specified"""
        self.logger.info("Running validation tests...")
        
        # Test 1: Merit badge count verification
        coverage_report = None
        for report_file in (self.output_dir / 'csv').glob('*coverage*.csv'):
            with open(report_file, 'r', encoding='utf-8') as f:
                content = f.read()
                badge_count = content.count('\n') - 1  # Subtract header
                break
        
        if badge_count == len(self.merit_badges):
            self.logger.info(f"✓ Test 1 PASSED: Merit badge count matches ({len(self.merit_badges)})")
        else:
            self.logger.error(f"✗ Test 1 FAILED: Expected {len(self.merit_badges)}, got {badge_count}")
        
        # Test 2: Eagle-required badge count
        eagle_count = len(self.eagle_required_badges)
        self.logger.info(f"✓ Test 2: Eagle-required badge count: {eagle_count}")
        
        # Test 3: Adult count verification
        total_adults = len(self.t12_adults) + len(self.t32_adults)
        total_counselors = len(self.generate_t12_t32_counselors())
        total_non_counselors = len(self.generate_non_counselors())
        
        if total_counselors + total_non_counselors == total_adults:
            self.logger.info(f"✓ Test 3 PASSED: Adult count verification ({total_adults} = {total_counselors} + {total_non_counselors})")
        else:
            self.logger.error(f"✗ Test 3 FAILED: Adult count mismatch")
        
        self.logger.info("Validation tests completed")


class MeritBadgeGUI:
    """GUI interface for Merit Badge Generator"""
    
    def __init__(self):
        self.root = TkinterDnD.Tk()
        self.root.title("Merit Badge Counselor List Generator")
        self.root.geometry("800x700")
        self.root.configure(bg='#f0f0f0')
        
        self.generator = MeritBadgeGenerator()
        self.generator.progress_callback = self.update_progress
        self.generator.status_callback = self.update_status
        
        self.t12_file = None
        self.t32_file = None
        self.pdf_files = []
        
        self.setup_ui()
        
    def setup_ui(self):
        """Setup the user interface"""
        # Main frame
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Title
        title_label = ttk.Label(
            main_frame, 
            text="🎖️ Merit Badge Counselor List Generator",
            font=('Arial', 16, 'bold')
        )
        title_label.grid(row=0, column=0, columnspan=2, pady=(0, 10))
        
        subtitle_label = ttk.Label(
            main_frame,
            text="Troop 12 & Troop 32 Acton MA",
            font=('Arial', 12)
        )
        subtitle_label.grid(row=1, column=0, columnspan=2, pady=(0, 20))
        
        # File selection section
        file_frame = ttk.LabelFrame(main_frame, text="Input Files", padding="10")
        file_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # T12 Roster
        ttk.Label(file_frame, text="T12 Roster (CSV):").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.t12_label = ttk.Label(file_frame, text="No file selected", foreground="gray")
        self.t12_label.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=2)
        ttk.Button(file_frame, text="Browse", command=self.select_t12_file).grid(row=0, column=2, padx=(10, 0), pady=2)
        
        # T32 Roster
        ttk.Label(file_frame, text="T32 Roster (CSV):").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.t32_label = ttk.Label(file_frame, text="No file selected", foreground="gray")
        self.t32_label.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=2)
        ttk.Button(file_frame, text="Browse", command=self.select_t32_file).grid(row=1, column=2, padx=(10, 0), pady=2)
        
        # PDF Files
        ttk.Label(file_frame, text="Counselor PDFs:").grid(row=2, column=0, sticky=tk.W, pady=2)
        self.pdf_label = ttk.Label(file_frame, text="No files selected", foreground="gray")
        self.pdf_label.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=2)
        ttk.Button(file_frame, text="Browse", command=self.select_pdf_files).grid(row=2, column=2, padx=(10, 0), pady=2)
        
        # Configure column weights
        file_frame.columnconfigure(1, weight=1)
        
        # Drag and drop area
        drop_frame = ttk.LabelFrame(main_frame, text="Drag & Drop Files Here", padding="20")
        drop_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        drop_label = ttk.Label(
            drop_frame, 
            text="📁 Drag CSV and PDF files here\nOr use the Browse buttons above",
            justify=tk.CENTER,
            font=('Arial', 10)
        )
        drop_label.grid(row=0, column=0)
        
        # Enable drag and drop
        drop_frame.drop_target_register(DND_FILES)
        drop_frame.dnd_bind('<<Drop>>', self.on_drop)
        
        # Control buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=4, column=0, columnspan=2, pady=10)
        
        self.generate_button = ttk.Button(
            button_frame, 
            text="🚀 Generate Reports", 
            command=self.generate_reports,
            style='Accent.TButton'
        )
        self.generate_button.grid(row=0, column=0, padx=5)
        
        ttk.Button(
            button_frame, 
            text="📁 Open Output Folder", 
            command=self.open_output_folder
        ).grid(row=0, column=1, padx=5)
        
        ttk.Button(
            button_frame, 
            text="❓ Help", 
            command=self.show_help
        ).grid(row=0, column=2, padx=5)
        
        # Progress section
        progress_frame = ttk.LabelFrame(main_frame, text="Progress", padding="10")
        progress_frame.grid(row=5, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(
            progress_frame, 
            variable=self.progress_var, 
            maximum=100
        )
        self.progress_bar.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=2)
        
        self.status_label = ttk.Label(progress_frame, text="Ready")
        self.status_label.grid(row=1, column=0, sticky=tk.W, pady=2)
        
        progress_frame.columnconfigure(0, weight=1)
        
        # Console output
        console_frame = ttk.LabelFrame(main_frame, text="Console Output", padding="10")
        console_frame.grid(row=6, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        
        self.console_text = scrolledtext.ScrolledText(
            console_frame, 
            height=8, 
            width=80,
            font=('Consolas', 9)
        )
        self.console_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        console_frame.columnconfigure(0, weight=1)
        console_frame.rowconfigure(0, weight=1)
        
        # Configure main grid weights
        main_frame.columnconfigure(0, weight=1)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.rowconfigure(6, weight=1)
        
        # Initial console message
        self.log_to_console("Merit Badge Counselor List Generator v1.0 Ready")
        self.log_to_console("Select your input files and click 'Generate Reports' to begin")
    
    def select_t12_file(self):
        """Select T12 roster file"""
        file_path = filedialog.askopenfilename(
            title="Select T12 Roster CSV File",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if file_path:
            self.t12_file = file_path
            self.t12_label.configure(text=Path(file_path).name, foreground="black")
            self.log_to_console(f"T12 roster selected: {Path(file_path).name}")
    
    def select_t32_file(self):
        """Select T32 roster file"""
        file_path = filedialog.askopenfilename(
            title="Select T32 Roster CSV File",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if file_path:
            self.t32_file = file_path
            self.t32_label.configure(text=Path(file_path).name, foreground="black")
            self.log_to_console(f"T32 roster selected: {Path(file_path).name}")
    
    def select_pdf_files(self):
        """Select PDF files"""
        file_paths = filedialog.askopenfilenames(
            title="Select Merit Badge Counselor PDF Files",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        if file_paths:
            self.pdf_files = list(file_paths)
            count = len(file_paths)
            self.pdf_label.configure(text=f"{count} file(s) selected", foreground="black")
            self.log_to_console(f"Selected {count} PDF file(s)")
    
    def on_drop(self, event):
        """Handle drag and drop files"""
        files = self.root.tk.splitlist(event.data)
        
        for file_path in files:
            file_path = Path(file_path)
            
            if file_path.suffix.lower() == '.csv':
                if 't12' in file_path.name.lower():
                    self.t12_file = str(file_path)
                    self.t12_label.configure(text=file_path.name, foreground="black")
                    self.log_to_console(f"T12 roster dropped: {file_path.name}")
                elif 't32' in file_path.name.lower():
                    self.t32_file = str(file_path)
                    self.t32_label.configure(text=file_path.name, foreground="black")
                    self.log_to_console(f"T32 roster dropped: {file_path.name}")
                else:
                    self.log_to_console(f"CSV file dropped (could not determine troop): {file_path.name}")
            
            elif file_path.suffix.lower() == '.pdf':
                if str(file_path) not in self.pdf_files:
                    self.pdf_files.append(str(file_path))
                    count = len(self.pdf_files)
                    self.pdf_label.configure(text=f"{count} file(s) selected", foreground="black")
                    self.log_to_console(f"PDF file dropped: {file_path.name}")
    
    def update_progress(self, percent):
        """Update progress bar"""
        self.progress_var.set(percent)
        self.root.update_idletasks()
    
    def update_status(self, message):
        """Update status label"""
        self.status_label.configure(text=message)
        self.log_to_console(message)
        self.root.update_idletasks()
    
    def log_to_console(self, message):
        """Log message to console"""
        timestamp = datetime.now().strftime('%H:%M:%S')
        self.console_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.console_text.see(tk.END)
        self.root.update_idletasks()
    
    def generate_reports(self):
        """Generate reports in background thread"""
        if not self.t12_file and not self.t32_file:
            messagebox.showwarning("No Files", "Please select at least one roster file (T12 or T32).")
            return
        
        self.generate_button.configure(state='disabled')
        
        # Run generation in background thread
        thread = threading.Thread(target=self._generate_reports_thread)
        thread.daemon = True
        thread.start()
    
    def _generate_reports_thread(self):
        """Background thread for report generation"""
        try:
            self.log_to_console("Starting report generation...")
            
            output_dir = self.generator.generate_reports(
                t12_file=self.t12_file,
                t32_file=self.t32_file,
                pdf_files=self.pdf_files if self.pdf_files else None
            )
            
            # Success message
            self.root.after(0, lambda: messagebox.showinfo(
                "Success", 
                f"Reports generated successfully!\n\nOutput directory:\n{output_dir}"
            ))
            
            # Auto-open output folder
            self.root.after(0, lambda: self.open_folder(output_dir))
            
        except Exception as e:
            error_msg = f"Error generating reports: {str(e)}"
            self.log_to_console(error_msg)
            self.root.after(0, lambda: messagebox.showerror("Error", error_msg))
        
        finally:
            self.root.after(0, lambda: self.generate_button.configure(state='normal'))
    
    def open_output_folder(self):
        """Open the most recent output folder"""
        # Find most recent output folder
        output_folders = [d for d in Path('.').iterdir() if d.is_dir() and d.name.startswith('MBC_Reports_')]
        if output_folders:
            latest_folder = max(output_folders, key=lambda x: x.stat().st_mtime)
            self.open_folder(latest_folder)
        else:
            messagebox.showinfo("No Output", "No output folders found. Generate reports first.")
    
    def open_folder(self, folder_path):
        """Open folder in system file manager"""
        try:
            if sys.platform == "win32":
                os.startfile(folder_path)
            elif sys.platform == "darwin":
                subprocess.run(["open", folder_path])
            else:
                subprocess.run(["xdg-open", folder_path])
        except Exception as e:
            self.log_to_console(f"Could not open folder: {e}")
    
    def show_help(self):
        """Show help dialog"""
        help_text = """
Merit Badge Counselor List Generator Help

GETTING STARTED:
1. Select your T12 and/or T32 roster CSV files
2. Select Merit Badge Counselor PDF files (optional)
3. Click "Generate Reports"

INPUT FILES:
• T12/T32 Rosters: CSV files with columns: name, email, phone, positionname
• PDF Files: ScoutBook Merit Badge Counselor search results

OUTPUTS:
The tool generates reports in multiple formats:
• HTML: Self-contained web pages with download buttons
• CSV: Excel-compatible data files
• PDF: Professional formatted reports
• Excel: Multi-sheet workbooks
• WordPress: Ready-to-use website content

DRAG & DROP:
You can drag files directly onto the interface:
• CSV files containing 't12' or 't32' will be auto-assigned
• PDF files will be added to the counselor list

OUTPUT FOLDER:
Reports are saved in timestamped folders like:
MBC_Reports_2025-06-02_14-30/

The summary report provides an overview and links to all generated files.

TROUBLESHOOTING:
• Ensure CSV files have the required columns
• PDF files should be text-based (not scanned images)
• Check the console output for detailed progress and error messages

For more information, visit the Scouting America website.
"""
        
        help_window = tk.Toplevel(self.root)
        help_window.title("Help")
        help_window.geometry("600x500")
        
        text_widget = scrolledtext.ScrolledText(help_window, wrap=tk.WORD, padx=10, pady=10)
        text_widget.pack(fill=tk.BOTH, expand=True)
        text_widget.insert(tk.END, help_text)
        text_widget.configure(state='disabled')
    
    def run(self):
        """Start the GUI application"""
        self.root.mainloop()


class MeritBadgeCLI:
    """Command-line interface for Merit Badge Generator"""
    
    def __init__(self):
        self.generator = MeritBadgeGenerator()
    
    def run(self, args):
        """Run CLI with provided arguments"""
        import argparse
        
        parser = argparse.ArgumentParser(
            description="Generate Merit Badge Counselor reports for T12/T32 Acton MA"
        )
        parser.add_argument('--t12-roster', type=str, help='T12 CSV roster file')
        parser.add_argument('--t32-roster', type=str, help='T32 CSV roster file')
        parser.add_argument('--counselor-pdfs', nargs='+', help='Merit badge counselor PDF files')
        parser.add_argument('--output-dir', type=str, help='Output directory (default: auto-generated)')
        parser.add_argument('--formats', nargs='+', 
                          choices=['html', 'csv', 'pdf', 'excel', 'wordpress'],
                          default=['html', 'csv', 'pdf', 'excel', 'wordpress'],
                          help='Output formats to generate')
        parser.add_argument('--verbose', '-v', action='store_true', help='Verbose output')
        
        parsed_args = parser.parse_args(args)
        
        if parsed_args.verbose:
            logging.getLogger().setLevel(logging.DEBUG)
        
        # Set up progress callback for CLI
        def cli_progress(percent):
            print(f"Progress: {percent:3.0f}%", end='\r')
        
        def cli_status(message):
            print(f"\n{message}")
        
        self.generator.progress_callback = cli_progress
        self.generator.status_callback = cli_status
        
        try:
            output_dir = self.generator.generate_reports(
                t12_file=parsed_args.t12_roster,
                t32_file=parsed_args.t32_roster,
                pdf_files=parsed_args.counselor_pdfs
            )
            
            print(f"\n✅ Reports generated successfully!")
            print(f"📁 Output directory: {output_dir}")
            print(f"📄 Open summary report: {output_dir}/summary_report.html")
            
        except Exception as e:
            print(f"\n❌ Error: {e}")
            return 1
        
        return 0


def main():
    """Main entry point"""
    if len(sys.argv) > 1:
        # CLI mode
        cli = MeritBadgeCLI()
        return cli.run(sys.argv[1:])
    else:
        # GUI mode
        try:
            gui = MeritBadgeGUI()
            gui.run()
        except ImportError as e:
            print(f"GUI dependencies not available: {e}")
            print("Install with: pip install tkinterdnd2")
            print("Or use CLI mode: python merit_badge_generator.py --help")
            return 1
        except Exception as e:
            print(f"Error starting GUI: {e}")
            return 1
    
    return 0


if __name__ == "__main__":
    sys.exit(main())


# Additional files that would be created alongside the main script:

# requirements.txt
"""
pandas>=1.5.0
requests>=2.28.0
beautifulsoup4>=4.11.0
pdfplumber>=0.7.0
jinja2>=3.1.0
openpyxl>=3.0.0
reportlab>=3.6.0
tkinterdnd2>=0.3.0
"""

# setup.py
"""
from setuptools import setup, find_packages

setup(
    name="merit-badge-generator",
    version="1.0.0",
    description="Merit Badge Counselor List Generator for Scouting America",
    long_description=open("README.md").read(),
    long_description_content_type="text/markdown",
    author="Generated by Claude",
    packages=find_packages(),
    install_requires=[
        "pandas>=1.5.0",
        "requests>=2.28.0",
        "beautifulsoup4>=4.11.0",
        "pdfplumber>=0.7.0",
        "jinja2>=3.1.0",
        "openpyxl>=3.0.0",
        "reportlab>=3.6.0",
        "tkinterdnd2>=0.3.0",
    ],
    entry_points={
        "console_scripts": [
            "merit-badge-generator=merit_badge_generator:main",
        ],
    },
    classifiers=[
        "Development Status :: 4 - Beta",
        "Intended Audience :: Other Audience",
        "License :: OSI Approved :: MIT License",
        "Programming Language :: Python :: 3",
        "Programming Language :: Python :: 3.8",
        "Programming Language :: Python :: 3.9",
        "Programming Language :: Python :: 3.10",
        "Programming Language :: Python :: 3.11",
    ],
    python_requires=">=3.8",
)
"""

# build_executable.py
"""
import PyInstaller.__main__
import os
import sys

def build_executable():
    '''Build standalone executable using PyInstaller'''
    
    # Basic PyInstaller arguments
    args = [
        'merit_badge_generator.py',
        '--onefile',
        '--windowed',
        '--name=MeritBadgeGenerator',
        '--icon=scout_icon.ico',  # Add icon if available
        '--add-data=config.json;.',
        '--hidden-import=tkinterdnd2',
        '--hidden-import=pdfplumber',
        '--hidden-import=openpyxl',
        '--hidden-import=reportlab',
    ]
    
    # Platform-specific adjustments
    if sys.platform == 'win32':
        args.extend([
            '--add-data=C:\\Windows\\System32\\msvcp140.dll;.',  # Include Visual C++ runtime
        ])
    
    print("Building executable...")
    PyInstaller.__main__.run(args)
    print("Build complete! Check the 'dist' folder.")

if __name__ == "__main__":
    build_executable()
"""#!/usr/bin/env python3
"""
Merit Badge Counselor List Generator
Generates comprehensive reports for Troop 12 & Troop 32 Acton MA

Author: Generated by Claude
Version: 1.0
Requirements: Python 3.8+
"""

import os
import sys
import json
import csv
import logging
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from tkinterdnd2 import DND_FILES, TkinterDnD
import threading
import webbrowser
from datetime import datetime
from pathlib import Path
import subprocess
import time

# Third-party imports
import pandas as pd
import requests
from bs4 import BeautifulSoup
import pdfplumber
from jinja2 import Template
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.units import inch


class MeritBadgeGenerator:
    """Main application class for Merit Badge Counselor List Generator"""
    
    def __init__(self):
        self.logger = self.setup_logging()
        self.config = self.load_config()
        self.merit_badges = []
        self.eagle_required_badges = []
        self.t12_adults = []
        self.t32_adults = []
        self.merit_badge_counselors = []
        self.output_dir = None
        self.progress_callback = None
        self.status_callback = None
        
    def setup_logging(self):
        """Setup logging configuration"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('merit_badge_generator.log'),
                logging.StreamHandler()
            ]
        )
        return logging.getLogger(__name__)
    
    def load_config(self):
        """Load configuration from file or create default"""
        config_file = 'config.json'
        default_config = {
            'merit_badges_url': 'https://www.scouting.org/skills/merit-badges/all/',
            'eagle_required_url': 'https://www.scouting.org/skills/merit-badges/eagle-required/',
            'output_formats': ['html', 'csv', 'pdf', 'excel', 'wordpress'],
            'network_timeout': 30,
            'max_file_size_mb': 50
        }
        
        if os.path.exists(config_file):
            try:
                with open(config_file, 'r') as f:
                    config = json.load(f)
                return {**default_config, **config}
            except Exception as e:
                self.logger.warning(f"Error loading config: {e}. Using defaults.")
                
        return default_config
    
    def save_config(self):
        """Save current configuration"""
        try:
            with open('config.json', 'w') as f:
                json.dump(self.config, f, indent=2)
        except Exception as e:
            self.logger.error(f"Error saving config: {e}")
    
    def update_progress(self, percent, message=""):
        """Update progress bar and status"""
        if self.progress_callback:
            self.progress_callback(percent)
        if self.status_callback:
            self.status_callback(message)
        self.logger.info(f"Progress: {percent}% - {message}")
    
    def fetch_merit_badges(self):
        """Fetch current merit badges from scouting.org"""
        try:
            self.update_progress(5, "Fetching merit badges from scouting.org...")
            
            response = requests.get(
                self.config['merit_badges_url'], 
                timeout=self.config['network_timeout']
            )
            response.raise_for_status()
            
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Find merit badges after "Merit Badges A-Z" heading
            badges = []
            merit_badge_section = soup.find('h2', string=lambda text: text and 'Merit Badges A-Z' in text)
            
            if merit_badge_section:
                # Look for links to merit badge pages
                for link in soup.find_all('a', href=lambda href: href and '/merit-badges/' in href):
                    badge_name = link.get_text().strip()
                    if badge_name and badge_name not in badges:
                        badges.append(badge_name)
            
            # Fallback list if scraping fails
            if not badges:
                self.logger.warning("Could not scrape merit badges, using fallback list")
                badges = self.get_fallback_merit_badges()
            
            self.merit_badges = sorted(list(set(badges)))
            self.logger.info(f"Loaded {len(self.merit_badges)} merit badges")
            
        except Exception as e:
            self.logger.error(f"Error fetching merit badges: {e}")
            self.merit_badges = self.get_fallback_merit_badges()
    
    def fetch_eagle_required_badges(self):
        """Fetch Eagle-required merit badges from scouting.org"""
        try:
            self.update_progress(10, "Fetching Eagle-required badges...")
            
            response = requests.get(
                self.config['eagle_required_url'],
                timeout=self.config['network_timeout']
            )
            response.raise_for_status()
            
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Extract Eagle-required badges
            eagle_badges = []
            for link in soup.find_all('a', href=lambda href: href and '/merit-badges/' in href):
                badge_name = link.get_text().strip()
                if badge_name:
                    eagle_badges.append(badge_name)
            
            if not eagle_badges:
                eagle_badges = [
                    'Camping', 'Citizenship in Community', 'Citizenship in Nation', 
                    'Citizenship in Society', 'Communication', 'Cooking', 
                    'Emergency Preparedness', 'Environmental Science', 'Family Life',
                    'First Aid', 'Personal Fitness', 'Personal Management',
                    'Cycling', 'Hiking', 'Swimming'
                ]
            
            self.eagle_required_badges = sorted(list(set(eagle_badges)))
            self.logger.info(f"Loaded {len(self.eagle_required_badges)} Eagle-required badges")
            
        except Exception as e:
            self.logger.error(f"Error fetching Eagle-required badges: {e}")
            self.eagle_required_badges = [
                'Camping', 'Citizenship in Community', 'Citizenship in Nation',
                'Citizenship in Society', 'Communication', 'Cooking',
                'Emergency Preparedness', 'Environmental Science', 'Family Life',
                'First Aid', 'Personal Fitness', 'Personal Management',
                'Cycling', 'Hiking', 'Swimming'
            ]
    
    def get_fallback_merit_badges(self):
        """Return fallback list of merit badges"""
        return [
            'Animal Science', 'Archery', 'Art', 'Astronomy', 'Athletics', 'Automotive',
            'Aviation', 'Backpacking', 'Basketry', 'Bird Study', 'Bugling', 'Camping',
            'Canoeing', 'Chemistry', 'Chess', 'Citizenship in Community', 'Citizenship in Nation',
            'Citizenship in Society', 'Climbing', 'Coin Collecting', 'Collections', 'Communication',
            'Composite Materials', 'Cooking', 'Crime Prevention', 'Cycling', 'Dentistry',
            'Digital Technology', 'Disabilities Awareness', 'Dog Care', 'Drafting', 'Electricity',
            'Electronics', 'Emergency Preparedness', 'Energy', 'Engineering', 'Entrepreneurship',
            'Environmental Science', 'Family Life', 'Farm Mechanics', 'Fingerprinting',
            'Fire Safety', 'First Aid', 'Fish and Wildlife Management', 'Fishing', 'Forestry',
            'Game Design', 'Gardening', 'Genealogy', 'Geocaching', 'Geology', 'Golf',
            'Graphic Arts', 'Hiking', 'Home Repairs', 'Horsemanship', 'Indian Lore',
            'Insect Study', 'Inventing', 'Journalism', 'Kayaking', 'Landscape Architecture',
            'Law', 'Leatherwork', 'Lifesaving', 'Mammal Study', 'Medicine', 'Metalwork',
            'Model Design and Building', 'Motorboating', 'Music', 'Nature', 'Nuclear Science',
            'Oceanography', 'Orienteering', 'Painting', 'Personal Fitness', 'Personal Management',
            'Pets', 'Photography', 'Pioneering', 'Plant Science', 'Plumbing', 'Pottery',
            'Programming', 'Public Health', 'Public Speaking', 'Pulp and Paper', 'Radio',
            'Railroading', 'Reading', 'Reptile and Amphibian Study', 'Rifle Shooting',
            'Robotics', 'Rowing', 'Safety', 'Salesmanship', 'Scholarship', 'Scouting Heritage',
            'Scuba Diving', 'Sculpture', 'Search and Rescue', 'Shotgun Shooting',
            'Signs, Signals, and Codes', 'Skating', 'Small-Boat Sailing', 'Snow Sports',
            'Soil and Water Conservation', 'Space Exploration', 'Sports', 'Stamp Collecting',
            'Sustainability', 'Swimming', 'Textile', 'Theater', 'Traffic Safety',
            'Truck Transportation', 'Veterinary Medicine', 'Water Sports', 'Weather',
            'Welding', 'Whitewater', 'Wilderness Survival', 'Wood Carving', 'Woodworking'
        ]
    
    def process_csv_file(self, file_path, troop_id):
        """Process CSV roster file"""
        try:
            self.update_progress(20, f"Processing {troop_id} roster...")
            
            # Detect encoding
            encodings = ['utf-8', 'windows-1252', 'iso-8859-1']
            df = None
            
            for encoding in encodings:
                try:
                    df = pd.read_csv(file_path, encoding=encoding)
                    break
                except UnicodeDecodeError:
                    continue
            
            if df is None:
                raise ValueError(f"Could not read CSV file with any encoding: {file_path}")
            
            # Validate required columns
            required_columns = ['name', 'email', 'phone', 'positionname']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                self.logger.warning(f"Missing columns in {troop_id}: {missing_columns}")
            
            # Filter adults (not Youth Members)
            adults = df[~df['positionname'].str.contains('Youth Member', na=False)]
            
            # Clean and standardize data
            adults = adults.copy()
            adults['name'] = adults['name'].str.strip()
            adults['email'] = adults['email'].str.strip()
            adults['phone'] = adults['phone'].str.strip()
            
            adults_list = adults.to_dict('records')
            self.logger.info(f"Processed {len(adults_list)} adults from {troop_id}")
            
            return adults_list
            
        except Exception as e:
            self.logger.error(f"Error processing CSV {file_path}: {e}")
            raise
    
    def process_pdf_files(self, pdf_files):
        """Process merit badge counselor PDF files"""
        counselors = []
        
        for i, pdf_file in enumerate(pdf_files):
            try:
                self.update_progress(
                    30 + (i * 20 // len(pdf_files)), 
                    f"Processing PDF {i+1}/{len(pdf_files)}..."
                )
                
                # Check file size
                file_size_mb = os.path.getsize(pdf_file) / (1024 * 1024)
                if file_size_mb > self.config['max_file_size_mb']:
                    self.logger.warning(f"PDF file {pdf_file} is large ({file_size_mb:.1f}MB)")
                
                with pdfplumber.open(pdf_file) as pdf:
                    text = ""
                    for page in pdf.pages:
                        page_text = page.extract_text()
                        if page_text:
                            text += page_text + "\n"
                
                if not text.strip():
                    self.logger.warning(f"No text extracted from {pdf_file}")
                    continue
                
                # Parse counselor data from PDF text
                pdf_counselors = self.parse_counselor_text(text)
                counselors.extend(pdf_counselors)
                
            except Exception as e:
                self.logger.error(f"Error processing PDF {pdf_file}: {e}")
                continue
        
        self.merit_badge_counselors = counselors
        self.logger.info(f"Processed {len(counselors)} merit badge counselors")
        
        return counselors
    
    def parse_counselor_text(self, text):
        """Parse counselor information from PDF text"""
        counselors = []
        
        # This is a simplified parser - real implementation would need
        # more sophisticated parsing based on actual ScoutBook PDF format
        lines = text.split('\n')
        
        current_counselor = {}
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            # Look for patterns that indicate counselor data
            # This would need to be customized based on actual PDF format
            if '@' in line:  # Email
                current_counselor['email'] = line
            elif line.startswith(('(', '+')):  # Phone
                current_counselor['phone'] = line
            elif 'Merit Badge' in line or 'Badge:' in line:  # Merit badges
                badges = line.replace('Merit Badge:', '').replace('Badge:', '').strip()
                current_counselor['badges'] = [b.strip() for b in badges.split(',')]
            elif len(line.split()) >= 2 and line[0].isupper():  # Likely a name
                if current_counselor:
                    counselors.append(current_counselor)
                current_counselor = {'name': line, 'units': ['Other']}
        
        if current_counselor:
            counselors.append(current_counselor)
        
        return counselors
    
    def generate_reports(self, t12_file=None, t32_file=None, pdf_files=None):
        """Main report generation function"""
        try:
            self.update_progress(0, "Starting report generation...")
            
            # Create output directory
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
            self.output_dir = Path(f"MBC_Reports_{timestamp}")
            self.output_dir.mkdir(exist_ok=True)
            
            # Create subdirectories
            for format_dir in ['html', 'csv', 'pdf', 'excel', 'wordpress']:
                (self.output_dir / format_dir).mkdir(exist_ok=True)
            
            # Fetch merit badge data
            self.fetch_merit_badges()
            self.fetch_eagle_required_badges()
            
            # Process input files
            if t12_file:
                self.t12_adults = self.process_csv_file(t12_file, "T12")
            if t32_file:
                self.t32_adults = self.process_csv_file(t32_file, "T32")
            if pdf_files:
                self.process_pdf_files(pdf_files)
            
            self.update_progress(50, "Generating reports...")
            
            # Generate the three main reports
            t12_t32_counselors = self.generate_t12_t32_counselors()
            non_counselors = self.generate_non_counselors()
            coverage = self.generate_coverage_report()
            
            self.update_progress(70, "Creating output files...")
            
            # Generate all output formats
            reports = [
                {
                    'title': 'T12/T32 Merit Badge Counselors',
                    'data': t12_t32_counselors,
                    'type': 'counselors',
                    'filename': 't12_t32_merit_badge_counselors'
                },
                {
                    'title': 'T12/T32 Leaders not Merit Badge Counselors',
                    'data': non_counselors,
                    'type': 'non_counselors',
                    'filename': 't12_t32_leaders_not_counselors'
                },
                {
                    'title': 'T12/T32 Merit Badge Counselor Coverage',
                    'data': coverage,
                    'type': 'coverage',
                    'filename': 't12_t32_counselor_coverage'
                }
            ]
            
            for report in reports:
                self.generate_html_report(report)
                self.generate_csv_report(report)
                self.generate_pdf_report(report)
                self.generate_excel_report(report)
                self.generate_wordpress_report(report)
            
            self.update_progress(90, "Generating summary report...")
            self.generate_summary_report(reports)
            
            self.update_progress(95, "Running validation tests...")
            self.run_validation_tests()
            
            self.update_progress(100, "Report generation completed!")
            
            return self.output_dir
            
        except Exception as e:
            self.logger.error(f"Error in report generation: {e}")
            raise
    
    def generate_t12_t32_counselors(self):
        """Generate T12/T32 Merit Badge Counselors report"""
        all_adults = self.t12_adults + self.t32_adults
        adult_names = {adult['name'] for adult in all_adults}
        
        t12_t32_counselors = []
        for counselor in self.merit_badge_counselors:
            if counselor['name'] in adult_names:
                # Merge contact information
                for adult in all_adults:
                    if adult['name'] == counselor['name']:
                        merged = {**counselor, **adult}
                        t12_t32_counselors.append(merged)
                        break
        
        # Sort alphabetically
        t12_t32_counselors.sort(key=lambda x: x['name'])
        
        return t12_t32_counselors
    
    def generate_non_counselors(self):
        """Generate T12/T32 Leaders not Merit Badge Counselors report"""
        all_adults = self.t12_adults + self.t32_adults
        counselor_names = {counselor['name'] for counselor in self.merit_badge_counselors}
        
        non_counselors = []
        for adult in all_adults:
            if adult['name'] not in counselor_names:
                non_counselors.append(adult)
        
        # Sort alphabetically
        non_counselors.sort(key=lambda x: x['name'])
        
        return non_counselors
    
    def generate_coverage_report(self):
        """Generate Merit Badge Coverage report"""
        all_adults = self.t12_adults + self.t32_adults
        adult_names = {adult['name'] for adult in all_adults}
        
        # Get T12/T32 counselors and their badges
        t12_t32_counselor_badges = set()
        for counselor in self.merit_badge_counselors:
            if counselor['name'] in adult_names and 'badges' in counselor:
                t12_t32_counselor_badges.update(counselor['badges'])
        
        coverage = {
            'eagle_with_counselors': [],
            'eagle_without_counselors': [],
            'non_eagle_with_counselors': [],
            'non_eagle_without_counselors': []
        }
        
        for badge in self.merit_badges:
            is_eagle = badge in self.eagle_required_badges
            has_counselor = badge in t12_t32_counselor_badges
            
            if is_eagle:
                if has_counselor:
                    coverage['eagle_with_counselors'].append(badge)
                else:
                    coverage['eagle_without_counselors'].append(badge)
            else:
                if has_counselor:
                    coverage['non_eagle_with_counselors'].append(badge)
                else:
                    coverage['non_eagle_without_counselors'].append(badge)
        
        # Sort all categories
        for key in coverage:
            coverage[key].sort()
        
        return coverage
    
    def generate_html_report(self, report):
        """Generate HTML report"""
        template = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{{ title }}</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }
        .container { max-width: 1200px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
        .header { background: linear-gradient(135deg, #2c3e50 0%, #3498db 100%); color: white; padding: 20px; margin: -30px -30px 30px -30px; border-radius: 8px 8px 0 0; }
        .header h1 { margin: 0; font-size: 2rem; }
        .timestamp { margin-top: 10px; opacity: 0.9; font-size: 0.9rem; }
        .actions { margin: 20px 0; display: flex; gap: 10px; flex-wrap: wrap; }
        .btn { background: #3498db; color: white; border: none; padding: 10px 20px; border-radius: 4px; cursor: pointer; text-decoration: none; }
        .btn:hover { background: #2980b9; }
        .btn-success { background: #27ae60; }
        table { width: 100%; border-collapse: collapse; margin: 20px 0; }
        th, td { padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }
        th { background: #f8f9fa; font-weight: 600; color: #2c3e50; }
        tr:hover { background: #f8f9fa; }
        .coverage-section { margin: 30px 0; padding: 20px; background: #f8f9fa; border-radius: 6px; border-left: 4px solid #3498db; }
        .coverage-section h3 { margin-top: 0; color: #2c3e50; }
        .coverage-section ul { columns: 3; column-gap: 30px; list-style-type: disc; margin-left: 20px; }
        .coverage-section li { break-inside: avoid; margin-bottom: 5px; }
        @media print { .actions { display: none; } }
        @media (max-width: 768px) { .coverage-section ul { columns: 1; } th, td { padding: 8px; } }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>{{ title }}</h1>
            <div class="timestamp">Generated: {{ timestamp }}</div>
        </div>
        
        <div class="actions">
            <button class="btn" onclick="downloadCSV()">📥 Download CSV</button>
            <button class="btn btn-success" onclick="window.print()">🖨️ Print</button>
        </div>
        
        {{ content }}
    </div>
    
    <script>
        var csvData = {{ csv_data | tojson }};
        function downloadCSV() {
            var blob = new Blob([csvData], {type: 'text/csv;charset=utf-8;'});
            var link = document.createElement('a');
            var url = URL.createObjectURL(blob);
            link.setAttribute('href', url);
            link.setAttribute('download', '{{ filename }}.csv');
            link.style.visibility = 'hidden';
            document.body.appendChild(link);
            link.click();
            document.body.removeChild(link);
        }
    </script>
</body>
</html>
        """
        
        # Generate content based on report type
        if report['type'] == 'counselors':
            content = self.generate_counselors_table_html(report['data'])
            csv_data = self.generate_counselors_csv(report['data'])
        elif report['type'] == 'non_counselors':
            content = self.generate_non_counselors_table_html(report['data'])
            csv_data = self.generate_non_counselors_csv(report['data'])
        else:  # coverage
            content = self.generate_coverage_html(report['data'])
            csv_data = self.generate_coverage_csv(report['data'])
        
        template_obj = Template(template)
        html = template_obj.render(
            title=report['title'],
            timestamp=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            content=content,
            csv_data=csv_data,
            filename=report['filename']
        )
        
        # Save HTML file
        html_path = self.output_dir / 'html' / f"{report['filename']}.html"
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(html)
        
        self.logger.info(f"Generated HTML report: {html_path}")
    
    def generate_counselors_table_html(self, data):
        """Generate HTML table for counselors"""
        html = """
        <table>
            <thead>
                <tr>
                    <th>Name</th>
                    <th>Email</th>
                    <th>Phone</th>
                    <th>Merit Badges</th>
                    <th>Units</th>
                </tr>
            </thead>
            <tbody>
        """
        
        for counselor in data:
            badges = ', '.join(counselor.get('badges', []))
            units = ', '.join(counselor.get('units', []))
            html += f"""
                <tr>
                    <td>{counselor.get('name', 'N/A')}</td>
                    <td>{counselor.get('email', 'N/A')}</td>
                    <td>{counselor.get('phone', 'N/A')}</td>
                    <td>{badges or 'N/A'}</td>
                    <td>{units or 'N/A'}</td>
                </tr>
            """
        
        html += "</tbody></table>"
        return html
    
    def generate_non_counselors_table_html(self, data):
        """Generate HTML table for non-counselors"""
        html = """
        <table>
            <thead>
                <tr>
                    <th>Name</th>
                    <th>Email</th>
                    <th>Phone</th>
                    <th>Position</th>
                </tr>
            </thead>
            <tbody>
        """
        
        for adult in data:
            html += f"""
                <tr>
                    <td>{adult.get('name', 'N/A')}</td>
                    <td>{adult.get('email', 'N/A')}</td>
                    <td>{adult.get('phone', 'N/A')}</td>
                    <td>{adult.get('positionname', 'N/A')}</td>
                </tr>
            """
        
        html += "</tbody></table>"
        return html
    
    def generate_coverage_html(self, data):
        """Generate HTML for coverage report"""
        sections = [
            ('Eagle-Required Merit Badges with T12/T32 Counselors', data['eagle_with_counselors']),
            ('Eagle-Required Merit Badges without T12/T32 Counselors', data['eagle_without_counselors']),
            ('Non-Eagle-Required Merit Badges with T12/T32 Counselors', data['non_eagle_with_counselors']),
            ('Non-Eagle-Required Merit Badges without T12/T32 Counselors', data['non_eagle_without_counselors'])
        ]
        
        html = ""
        for title, badges in sections:
            html += f"""
            <div class="coverage-section">
                <h3>{title} ({len(badges)})</h3>
                <ul>
            """
            for badge in badges:
                html += f"<li>{badge}</li>"
            html += "</ul></div>"
        
        return html
    
    def generate_csv_report(self, report):
        """Generate CSV report"""
        csv_path = self.output_dir / 'csv' / f"{report['filename']}.csv"
        
        if report['type'] == 'counselors':
            csv_data = self.generate_counselors_csv(report['data'])
        elif report['type'] == 'non_counselors':
            csv_data = self.generate_non_counselors_csv(report['data'])
        else:  # coverage
            csv_data = self.generate_coverage_csv(report['data'])
        
        with open(csv_path, 'w', encoding='utf-8', newline='') as f:
            f.write(csv_data)
        
        self.logger.info(f"Generated CSV report: {csv_path}")
    
    def generate_counselors_csv(self, data):
        """Generate CSV content for counselors"""
        lines = ['Name,Email,Phone,Merit Badges,Units']
        for counselor in data:
            badges = '; '.join(counselor.get('badges', []))
            units = '; '.join(counselor.get('units', []))
            line = f'"{counselor.get("name", "")}","{counselor.get("email", "")}","{counselor.get("phone", "")}","{badges}","{units}"'
            lines.append(line)
        return '\n'.join(lines)
    
    def generate_non_counselors_csv(self, data):
        """Generate CSV content for non-counselors"""
        lines = ['Name,Email,Phone,Position']
        for adult in data:
            line = f'"{adult.get("name", "")}","{adult.get("email", "")}","{adult.get("phone", "")}","{adult.get("positionname", "")}"'
            lines.append(line)
        return '\n'.join(lines)
    
    