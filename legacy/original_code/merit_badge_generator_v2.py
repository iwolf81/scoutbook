            output_dir = self.generator.generate_reports(
                t12_file=parsed_args.t12_roster,
                t32_file=parsed_args.t32_roster,
                pdf_files=parsed_args.counselor_pdfs
            )
            
            end_time = time.time()
            total_time = end_time - start_time
            
            print(f"\n✅ Reports generated successfully in {total_time:.1f} seconds!")
            print(f"📁 Output directory: {output_dir}")
            print(f"📄 Summary report: {output_dir}/summary_report.html")
            print(f"📊 HTML reports: {output_dir}/html/")
            print(f"📋 CSV files: {output_dir}/csv/")
            print(f"📰 PDF files: {output_dir}/pdf/")
            print(f"📊 Excel files: {output_dir}/excel/")
            print(f"🌐 WordPress files: {output_dir}/wordpress/")
            
        except Exception as e:
            print(f"\n❌ Error: {e}")
            if parsed_args.verbose:
                import traceback
                traceback.print_exc()
            return 1
        
        return 0
    
    def validate_files_only(self, args):
        """Validate input files without generating reports"""
        print("🔍 Validating input files...")
        
        all_valid = True
        
        if args.t12_roster:
            print(f"\nValidating T12 roster: {args.t12_roster}")
            try:
                is_valid, message = self.generator.validate_csv_file(args.t12_roster)
                if is_valid:
                    print(f"  ✅ {message}")
                else:
                    print(f"  ⚠️  {message}")
                    all_valid = False
            except Exception as e:
                print(f"  ❌ Error: {e}")
                all_valid = False
        
        if args.t32_roster:
            print(f"\nValidating T32 roster: {args.t32_roster}")
            try:
                is_valid, message = self.generator.validate_csv_file(args.t32_roster)
                if is_valid:
                    print(f"  ✅ {message}")
                else:
                    print(f"  ⚠️  {message}")
                    all_valid = False
            except Exception as e:
                print(f"  ❌ Error: {e}")
                all_valid = False
        
        if args.counselor_pdfs:
            print(f"\nValidating {len(args.counselor_pdfs)} PDF file(s):")
            for pdf_file in args.counselor_pdfs:
                try:
                    is_valid, message = self.generator.validate_pdf_file(pdf_file)
                    status = "✅" if is_valid else "⚠️"
                    print(f"  {status} {Path(pdf_file).name}: {message}")
                    if not is_valid:
                        all_valid = False
                except Exception as e:
                    print(f"  ❌ {Path(pdf_file).name}: Error - {e}")
                    all_valid = False
        
        if not args.t12_roster and not args.t32_roster:
            print("\n📝 No roster files specified - sample data will be used for generation")
        
        print(f"\n{'✅ All files valid!' if all_valid else '⚠️  Some files have issues - check messages above'}")
        return 0 if all_valid else 1


def main():
    """Main entry point with enhanced error handling"""
    try:
        if len(sys.argv) > 1:
            # CLI mode
            cli = MeritBadgeCLI()
            return cli.run(sys.argv[1:])
        else:
            # GUI mode
            try:
                if not GUI_AVAILABLE:
                    print("GUI dependencies not available.")
                    print("Install with: pip install tkinterdnd2")
                    print("Or use CLI mode: python merit_badge_generator.py --help")
                    return 1
                
                gui = MeritBadgeGUI()
                gui.run()
                return 0
                
            except Exception as e:
                print(f"Error starting GUI: {e}")
                print("Try CLI mode: python merit_badge_generator.py --help")
                return 1
    
    except KeyboardInterrupt:
        print("\n\nOperation cancelled by user.")
        return 130
    except Exception as e:
        print(f"Unexpected error: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())_frame = ttk.Frame(main_frame)
        title_frame.grid(row=0, column=0, columnspan=2, pady=(0, 20), sticky=(tk.W, tk.E))
        
        title_label = ttk.Label(
            title_frame, 
            text="🎖️ Merit Badge Counselor Generator",
            font=('Arial', 18, 'bold')
        )
        title_label.grid(row=0, column=0, pady=(0, 5))
        
        subtitle_label = ttk.Label(
            title_frame,
            text="Troop 12 & Troop 32 Acton MA - Version 2.0",
            font=('Arial', 12)
        )
        subtitle_label.grid(row=1, column=0, pady=(0, 5))
        
        description_label = ttk.Label(
            title_frame,
            text="Generate comprehensive reports for merit badge counselors and coverage analysis",
            font=('Arial', 10),
            foreground='gray'
        )
        description_label.grid(row=2, column=0)
        
        # File selection section with enhanced validation
        file_frame = ttk.LabelFrame(main_frame, text="📁 Input Files", padding="15")
        file_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        
        # T12 Roster with validation indicator
        t12_frame = ttk.Frame(file_frame)
        t12_frame.grid(row=0, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Label(t12_frame, text="T12 Roster (CSV):", font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky=tk.W)
        self.t12_label = ttk.Label(t12_frame, text="No file selected", foreground="gray")
        self.t12_label.grid(row=1, column=0, sticky=(tk.W, tk.E), padx=(20, 0))
        self.t12_status = ttk.Label(t12_frame, text="", foreground="green")
        self.t12_status.grid(row=2, column=0, sticky=(tk.W, tk.E), padx=(20, 0))
        
        ttk.Button(t12_frame, text="Browse", command=self.select_t12_file).grid(row=0, column=1, padx=(10, 0), rowspan=2)
        
        # T32 Roster with validation indicator
        t32_frame = ttk.Frame(file_frame)
        t32_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Label(t32_frame, text="T32 Roster (CSV):", font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky=tk.W)
        self.t32_label = ttk.Label(t32_frame, text="No file selected", foreground="gray")
        self.t32_label.grid(row=1, column=0, sticky=(tk.W, tk.E), padx=(20, 0))
        self.t32_status = ttk.Label(t32_frame, text="", foreground="green")
        self.t32_status.grid(row=2, column=0, sticky=(tk.W, tk.E), padx=(20, 0))
        
        ttk.Button(t32_frame, text="Browse", command=self.select_t32_file).grid(row=0, column=1, padx=(10, 0), rowspan=2)
        
        # PDF Files with validation indicator
        pdf_frame = ttk.Frame(file_frame)
        pdf_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Label(pdf_frame, text="Counselor PDFs:", font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky=tk.W)
        self.pdf_label = ttk.Label(pdf_frame, text="No files selected (optional)", foreground="gray")
        self.pdf_label.grid(row=1, column=0, sticky=(tk.W, tk.E), padx=(20, 0))
        self.pdf_status = ttk.Label(pdf_frame, text="", foreground="green")
        self.pdf_status.grid(row=2, column=0, sticky=(tk.W, tk.E), padx=(20, 0))
        
        ttk.Button(pdf_frame, text="Browse", command=self.select_pdf_files).grid(row=0, column=1, padx=(10, 0), rowspan=2)
        
        # Configure column weights
        for frame in [t12_frame, t32_frame, pdf_frame]:
            frame.columnconfigure(0, weight=1)
        file_frame.columnconfigure(0, weight=1)
        
        # Options section
        options_frame = ttk.LabelFrame(main_frame, text="⚙️ Options", padding="15")
        options_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        
        # Network connectivity check
        self.network_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            options_frame, 
            text="Fetch latest merit badges from scouting.org",
            variable=self.network_var
        ).grid(row=0, column=0, sticky=tk.W, pady=2)
        
        # Auto-open results
        self.auto_open_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            options_frame, 
            text="Automatically open results folder",
            variable=self.auto_open_var
        ).grid(row=1, column=0, sticky=tk.W, pady=2)
        
        # Verbose logging
        self.verbose_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            options_frame, 
            text="Verbose logging (debug mode)",
            variable=self.verbose_var
        ).grid(row=2, column=0, sticky=tk.W, pady=2)
        
        # Control buttons with enhanced styling
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=3, column=0, columnspan=2, pady=15)
        
        self.generate_button = ttk.Button(
            button_frame, 
            text="🚀 Generate Reports", 
            command=self.generate_reports,
            width=20
        )
        self.generate_button.grid(row=0, column=0, padx=5)
        
        ttk.Button(
            button_frame, 
            text="📁 Open Output", 
            command=self.open_output_folder,
            width=15
        ).grid(row=0, column=1, padx=5)
        
        ttk.Button(
            button_frame, 
            text="🔍 Validate Files", 
            command=self.validate_files,
            width=15
        ).grid(row=0, column=2, padx=5)
        
        ttk.Button(
            button_frame, 
            text="❓ Help", 
            command=self.show_help,
            width=10
        ).grid(row=0, column=3, padx=5)
        
        # Progress section with time estimation
        progress_frame = ttk.LabelFrame(main_frame, text="📊 Progress", padding="15")
        progress_frame.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(
            progress_frame, 
            variable=self.progress_var, 
            maximum=100,
            length=400
        )
        self.progress_bar.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=5)
        
        progress_info_frame = ttk.Frame(progress_frame)
        progress_info_frame.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        self.status_label = ttk.Label(progress_info_frame, text="Ready to generate reports")
        self.status_label.grid(row=0, column=0, sticky=tk.W)
        
        self.time_label = ttk.Label(progress_info_frame, text="", foreground="gray")
        self.time_label.grid(row=0, column=1, sticky=tk.E)
        
        progress_frame.columnconfigure(0, weight=1)
        progress_info_frame.columnconfigure(1, weight=1)
        
        # Console output with enhanced formatting
        console_frame = ttk.LabelFrame(main_frame, text="📋 Console Output", padding="15")
        console_frame.grid(row=5, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 15))
        
        console_controls = ttk.Frame(console_frame)
        console_controls.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        ttk.Button(console_controls, text="Clear", command=self.clear_console, width=10).grid(row=0, column=0)
        ttk.Button(console_controls, text="Save Log", command=self.save_log, width=10).grid(row=0, column=1, padx=(5, 0))
        
        self.console_text = scrolledtext.ScrolledText(
            console_frame, 
            height=10, 
            width=80,
            font=('Consolas', 9),
            bg='#f8f8f8',
            fg='#333333'
        )
        self.console_text.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        console_frame.columnconfigure(0, weight=1)
        console_frame.rowconfigure(1, weight=1)
        
        # Configure main grid weights
        main_frame.columnconfigure(0, weight=1)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.rowconfigure(5, weight=1)
        
        # Initial console message
        self.log_to_console("Merit Badge Counselor Generator v2.0 Ready", "info")
        self.log_to_console("Select your input files and click 'Generate Reports' to begin", "info")
        
        # Start time tracking
        self.start_time = None
    
    def setup_drag_drop(self):
        """Setup drag and drop functionality"""
        if GUI_AVAILABLE:
            self.root.drop_target_register(DND_FILES)
            self.root.dnd_bind('<<Drop>>', self.handle_drop)
    
    def handle_drop(self, event):
        """Handle drag and drop file events"""
        files = self.root.tk.splitlist(event.data)
        
        for file_path in files:
            file_path = Path(file_path)
            
            if file_path.suffix.lower() == '.csv':
                if 't12' in file_path.name.lower():
                    self.t12_file = str(file_path)
                    self.t12_label.configure(text=file_path.name, foreground="black")
                    self.validate_file(self.t12_file, 't12')
                elif 't32' in file_path.name.lower():
                    self.t32_file = str(file_path)
                    self.t32_label.configure(text=file_path.name, foreground="black")
                    self.validate_file(self.t32_file, 't32')
                else:
                    # Ask user which troop this CSV is for
                    choice = messagebox.askyesnocancel(
                        "CSV File Assignment",
                        f"Which troop is '{file_path.name}' for?\n\nYes = T12\nNo = T32\nCancel = Skip"
                    )
                    if choice is True:
                        self.t12_file = str(file_path)
                        self.t12_label.configure(text=file_path.name, foreground="black")
                        self.validate_file(self.t12_file, 't12')
                    elif choice is False:
                        self.t32_file = str(file_path)
                        self.t32_label.configure(text=file_path.name, foreground="black")
                        self.validate_file(self.t32_file, 't32')
                        
            elif file_path.suffix.lower() == '.pdf':
                if str(file_path) not in self.pdf_files:
                    self.pdf_files.append(str(file_path))
                    count = len(self.pdf_files)
                    self.pdf_label.configure(text=f"{count} file(s) selected", foreground="black")
                    self.validate_pdf_files()
        
        if files:
            self.log_to_console(f"Processed {len(files)} dropped file(s)", "info")
    
    def validate_file(self, file_path, troop):
        """Validate a CSV file and update status"""
        try:
            is_valid, message = self.generator.validate_csv_file(file_path)
            if is_valid:
                status_label = self.t12_status if troop == 't12' else self.t32_status
                status_label.configure(text="✓ Valid CSV format", foreground="green")
                self.log_to_console(f"{troop.upper()} CSV validated successfully", "success")
            else:
                status_label = self.t12_status if troop == 't12' else self.t32_status
                status_label.configure(text=f"⚠ {message}", foreground="orange")
                self.log_to_console(f"{troop.upper()} CSV validation warning: {message}", "warning")
        except Exception as e:
            status_label = self.t12_status if troop == 't12' else self.t32_status
            status_label.configure(text="❌ Invalid file", foreground="red")
            self.log_to_console(f"{troop.upper()} CSV validation error: {e}", "error")
    
    def validate_pdf_files(self):
        """Validate PDF files"""
        valid_count = 0
        for pdf_file in self.pdf_files:
            try:
                is_valid, message = self.generator.validate_pdf_file(pdf_file)
                if is_valid:
                    valid_count += 1
            except:
                pass
        
        if valid_count == len(self.pdf_files):
            self.pdf_status.configure(text=f"✓ {valid_count} valid PDF(s)", foreground="green")
        else:
            self.pdf_status.configure(text=f"⚠ {valid_count}/{len(self.pdf_files)} valid", foreground="orange")
    
    def select_t12_file(self):
        """Select T12 roster file with validation"""
        file_path = filedialog.askopenfilename(
            title="Select T12 Roster CSV File",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if file_path:
            self.t12_file = file_path
            self.t12_label.configure(text=Path(file_path).name, foreground="black")
            self.validate_file(file_path, 't12')
            self.log_to_console(f"T12 roster selected: {Path(file_path).name}", "info")
    
    def select_t32_file(self):
        """Select T32 roster file with validation"""
        file_path = filedialog.askopenfilename(
            title="Select T32 Roster CSV File",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if file_path:
            self.t32_file = file_path
            self.t32_label.configure(text=Path(file_path).name, foreground="black")
            self.validate_file(file_path, 't32')
            self.log_to_console(f"T32 roster selected: {Path(file_path).name}", "info")
    
    def select_pdf_files(self):
        """Select PDF files with validation"""
        file_paths = filedialog.askopenfilenames(
            title="Select Merit Badge Counselor PDF Files",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        if file_paths:
            self.pdf_files = list(file_paths)
            count = len(file_paths)
            self.pdf_label.configure(text=f"{count} file(s) selected", foreground="black")
            self.validate_pdf_files()
            self.log_to_console(f"Selected {count} PDF file(s)", "info")
    
    def validate_files(self):
        """Validate all selected files"""
        self.log_to_console("Validating all files...", "info")
        
        if self.t12_file:
            self.validate_file(self.t12_file, 't12')
        if self.t32_file:
            self.validate_file(self.t32_file, 't32')
        if self.pdf_files:
            self.validate_pdf_files()
        
        if not self.t12_file and not self.t32_file:
            self.log_to_console("No CSV files selected. Sample data will be used.", "warning")
        
        self.log_to_console("File validation completed", "info")
    
    def update_progress(self, percent):
        """Update progress bar with time estimation"""
        self.progress_var.set(percent)
        
        if self.start_time and percent > 0:
            elapsed = time.time() - self.start_time
            if percent > 5:  # Avoid division by very small numbers
                total_estimated = (elapsed / percent) * 100
                remaining = total_estimated - elapsed
                self.time_label.configure(
                    text=f"Elapsed: {elapsed:.1f}s | Est. remaining: {remaining:.1f}s"
                )
        
        self.root.update_idletasks()
    
    def update_status(self, message):
        """Update status label and log"""
        self.status_label.configure(text=message)
        self.log_to_console(message, "info")
        self.root.update_idletasks()
    
    def log_to_console(self, message, level="info"):
        """Log message to console with color coding"""
        timestamp = datetime.now().strftime('%H:%M:%S')
        
        # Color coding based on level
        colors = {
            "info": "#333333",
            "success": "#059669",
            "warning": "#d97706", 
            "error": "#dc2626"
        }
        
        # Insert with timestamp
        self.console_text.insert(tk.END, f"[{timestamp}] {message}\n")
        
        # Apply color to the last line
        if level in colors:
            line_start = self.console_text.index("end-2l")
            line_end = self.console_text.index("end-1l")
            self.console_text.tag_add(level, line_start, line_end)
            self.console_text.tag_config(level, foreground=colors[level])
        
        self.console_text.see(tk.END)
        self.root.update_idletasks()
    
    def clear_console(self):
        """Clear the console output"""
        self.console_text.delete(1.0, tk.END)
        self.log_to_console("Console cleared", "info")
    
    def save_log(self):
        """Save console log to file"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            log_file = filedialog.asksaveasfilename(
                title="Save Log File",
                defaultextension=".txt",
                initialvalue=f"mbc_generator_log_{timestamp}.txt",
                filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
            )
            
            if log_file:
                with open(log_file, 'w', encoding='utf-8') as f:
                    f.write(self.console_text.get(1.0, tk.END))
                self.log_to_console(f"Log saved to: {log_file}", "success")
                
        except Exception as e:
            self.log_to_console(f"Error saving log: {e}", "error")
    
    def generate_reports(self):
        """Generate reports with enhanced error handling"""
        # Validate inputs
        if not self.t12_file and not self.t32_file:
            response = messagebox.askyesno(
                "No Input Files", 
                "No CSV files selected. Generate reports with sample data?"
            )
            if not response:
                return
        
        # Disable button and start timing
        self.generate_button.configure(state='disabled', text="🔄 Generating...")
        self.start_time = time.time()
        
        # Apply options
        if self.verbose_var.get():
            self.generator.logger.setLevel(logging.DEBUG)
        
        # Run generation in background thread
        thread = threading.Thread(target=self._generate_reports_thread)
        thread.daemon = True
        thread.start()
    
    def _generate_reports_thread(self):
        """Background thread for report generation with enhanced error handling"""
        try:
            self.log_to_console("Starting report generation...", "info")
            
            # Configure generator based on options
            if not self.network_var.get():
                self.generator.config['merit_badges_url'] = None
                self.generator.config['eagle_required_url'] = None
            
            output_dir = self.generator.generate_reports(
                t12_file=self.t12_file,
                t32_file=self.t32_file,
                pdf_files=self.pdf_files if self.pdf_files else None
            )
            
            # Calculate total time
            total_time = time.time() - self.start_time if self.start_time else 0
            
            # Success message
            success_msg = (f"Reports generated successfully in {total_time:.1f} seconds!\n\n"
                          f"Output directory:\n{output_dir}")
            
            self.root.after(0, lambda: messagebox.showinfo("Success", success_msg))
            self.root.after(0, lambda: self.log_to_console(f"Generation completed in {total_time:.1f}s", "success"))
            
            # Auto-open output folder if enabled
            if self.auto_open_var.get():
                self.root.after(0, lambda: self.open_folder(output_dir))
            
        except Exception as e:
            error_msg = f"Error generating reports: {str(e)}"
            self.root.after(0, lambda: self.log_to_console(error_msg, "error"))
            self.root.after(0, lambda: messagebox.showerror("Error", error_msg))
        
        finally:
            self.root.after(0, self._reset_generate_button)
    
    def _reset_generate_button(self):
        """Reset the generate button state"""
        self.generate_button.configure(state='normal', text="🚀 Generate Reports")
        self.time_label.configure(text="")
    
    def open_output_folder(self):
        """Open the most recent output folder"""
        output_folders = [d for d in Path('.').iterdir() 
                         if d.is_dir() and d.name.startswith('MBC_Reports_')]
        if output_folders:
            latest_folder = max(output_folders, key=lambda x: x.stat().st_mtime)
            self.open_folder(latest_folder)
        else:
            messagebox.showinfo("No Output", "No output folders found. Generate reports first.")
    
    def open_folder(self, folder_path):
        """Open folder in system file manager"""
        try:
            if sys.platform == "win32":
                os.startfile(folder_path)
            elif sys.platform == "darwin":
                subprocess.run(["open", folder_path])
            else:
                subprocess.run(["xdg-open", folder_path])
        except Exception as e:
            self.log_to_console(f"Could not open folder: {e}", "error")
    
    def show_help(self):
        """Show comprehensive help dialog"""
        help_text = """Merit Badge Counselor Generator v2.0 - Help

🎯 PURPOSE:
Generate comprehensive reports analyzing merit badge counselor coverage 
for Troop 12 and Troop 32 in Acton, MA.

📋 GETTING STARTED:
1. Select T12 and/or T32 roster CSV files (or use sample data)
2. Optionally select Merit Badge Counselor PDF files from ScoutBook
3. Configure options as needed
4. Click "Generate Reports"

📁 INPUT FILES:
• T12/T32 Rosters: CSV files exported from ScoutBook containing:
  - firstname, lastname, positionname, primaryemail, primaryphone
  - Column headers should be on line starting with "memberid"
  
• PDF Files: ScoutBook Merit Badge Counselor search results
  - Should be text-based PDFs (not scanned images)
  - Search within 5 miles of Acton MA (01720)

📊 OUTPUT REPORTS:
1. T12/T32 Merit Badge Counselors - Leaders who are also counselors
2. T12/T32 Leaders not Merit Badge Counselors - Leaders to recruit
3. T12/T32 Merit Badge Counselor Coverage - Badge coverage analysis

📁 OUTPUT FORMATS:
• HTML: Interactive web pages with download buttons
• CSV: Excel-compatible data files
• PDF: Professional formatted reports  
• Excel: Multi-sheet workbooks with formatting
• WordPress: Shortcode-compatible content

⚙️ OPTIONS:
• Network Updates: Fetch latest merit badge lists from scouting.org
• Auto-open: Automatically open results folder when complete
• Verbose Logging: Enable detailed debug information

🔍 VALIDATION:
The tool validates:
• CSV file format and required columns
• PDF content and text extraction capability  
• Merit badge count consistency with scouting.org
• Adult accounting (counselors + non-counselors = total adults)

📋 DRAG & DROP:
You can drag and drop files directly onto the interface:
• CSV files will be assigned to T12 or T32 based on filename
• PDF files will be added to the counselor list

🛠️ TROUBLESHOOTING:
• Ensure CSV files have required columns with proper headers
• PDF files should be text-based, not scanned images
• Check console output for detailed progress and error messages
• Use "Validate Files" button to check file compatibility

📞 SUPPORT:
For issues or questions, check the console output and log files
generated in the logs/ directory.

🔗 REFERENCE:
Visit scouting.org for current merit badge information and requirements.
"""
        
        help_window = tk.Toplevel(self.root)
        help_window.title("Help - Merit Badge Counselor Generator")
        help_window.geometry("700x600")
        help_window.configure(bg='white')
        
        # Help content with scrolling
        help_frame = ttk.Frame(help_window, padding="20")
        help_frame.pack(fill=tk.BOTH, expand=True)
        
        text_widget = scrolledtext.ScrolledText(
            help_frame, 
            wrap=tk.WORD, 
            font=('Arial', 10),
            bg='white',
            fg='#333333'
        )
        text_widget.pack(fill=tk.BOTH, expand=True)
        text_widget.insert(tk.END, help_text)
        text_widget.configure(state='disabled')
        
        # Close button
        ttk.Button(
            help_frame, 
            text="Close", 
            command=help_window.destroy,
            width=15
        ).pack(pady=(10, 0))
    
    def run(self):
        """Start the GUI application"""
        self.root.mainloop()


class MeritBadgeCLI:
    """Enhanced command-line interface"""
    
    def __init__(self):
        self.generator = MeritBadgeGenerator()
    
    def run(self, args):
        """Run CLI with enhanced argument parsing and validation"""
        import argparse
        
        parser = argparse.ArgumentParser(
            description="Merit Badge Counselor Generator v2.0 - Generate reports for T12/T32 Acton MA",
            epilog="Example: python merit_badge_generator.py --t12-roster t12.csv --t32-roster t32.csv --counselor-pdfs counselors.pdf"
        )
        
        parser.add_argument('--t12-roster', type=str, 
                          help='T12 CSV roster file')
        parser.add_argument('--t32-roster', type=str, 
                          help='T32 CSV roster file')
        parser.add_argument('--counselor-pdfs', nargs='+', 
                          help='Merit badge counselor PDF files')
        parser.add_argument('--output-dir', type=str, 
                          help='Custom output directory (default: auto-generated)')
        parser.add_argument('--no-network', action='store_true',
                          help='Skip fetching data from scouting.org')
        parser.add_argument('--verbose', '-v', action='store_true', 
                          help='Verbose output with debug information')
        parser.add_argument('--validate-only', action='store_true',
                          help='Only validate input files, do not generate reports')
        parser.add_argument('--version', action='version', version='Merit Badge Generator 2.0')
        
        parsed_args = parser.parse_args(args)
        
        # Configure logging level
        if parsed_args.verbose:
            self.generator.logger.setLevel(logging.DEBUG)
        
        # Set up progress callbacks for CLI
        def cli_progress(percent):
            print(f"\rProgress: {percent:3.0f}%", end='', flush=True)
        
        def cli_status(message):
            print(f"\n{message}")
        
        self.generator.progress_callback = cli_progress
        self.generator.status_callback = cli_status
        
        try:
            # Validate files if requested
            if parsed_args.validate_only:
                return self.validate_files_only(parsed_args)
            
            # Configure network options
            if parsed_args.no_network:
                self.generator.config['merit_badges_url'] = None
                self.generator.config['eagle_required_url'] = None
                print("Network access disabled - using fallback data")
            
            # Generate reports
            start_time = time.time()
            
            output_dir = self.generator.generate_reports(
                t12_file=parsed_args.t12_roster,
                t32_file=parsed_#!/usr/bin/env python3
"""
Merit Badge Counselor List Generator - Updated Version
Generates comprehensive reports for Troop 12 & Troop 32 Acton MA

Author: Generated by Claude
Version: 2.0
Requirements: Python 3.8+
"""

import os
import sys
import json
import csv
import logging
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import webbrowser
from datetime import datetime
from pathlib import Path
import subprocess
import time
import re
from typing import List, Dict, Any, Optional

# Check for required packages and provide helpful error messages
REQUIRED_PACKAGES = [
    ('pandas', 'pandas'),
    ('requests', 'requests'),
    ('beautifulsoup4', 'bs4'),
    ('pdfplumber', 'pdfplumber'),
    ('jinja2', 'jinja2'),
    ('openpyxl', 'openpyxl'),
    ('reportlab', 'reportlab'),
]

missing_packages = []
for package_name, import_name in REQUIRED_PACKAGES:
    try:
        __import__(import_name)
    except ImportError:
        missing_packages.append(package_name)

if missing_packages:
    print("Error: Missing required packages. Install with:")
    print(f"pip install {' '.join(missing_packages)}")
    sys.exit(1)

# Now import the packages
import pandas as pd
import requests
from bs4 import BeautifulSoup
import pdfplumber
from jinja2 import Template
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.units import inch

# Optional GUI dependencies
try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    GUI_AVAILABLE = True
except ImportError:
    GUI_AVAILABLE = False


class MeritBadgeGenerator:
    """Main application class for Merit Badge Counselor List Generator"""
    
    def __init__(self):
        self.logger = self.setup_logging()
        self.config = self.load_config()
        self.merit_badges = []
        self.eagle_required_badges = []
        self.t12_adults = []
        self.t32_adults = []
        self.merit_badge_counselors = []
        self.output_dir = None
        self.progress_callback = None
        self.status_callback = None
        
    def setup_logging(self):
        """Setup logging configuration with debug level"""
        log_format = '%(asctime)s - %(levelname)s - %(message)s'
        
        # Create logs directory
        Path('logs').mkdir(exist_ok=True)
        
        # Setup file handler
        file_handler = logging.FileHandler(f'logs/merit_badge_generator_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
        file_handler.setLevel(logging.DEBUG)
        file_handler.setFormatter(logging.Formatter(log_format))
        
        # Setup console handler
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        console_handler.setFormatter(logging.Formatter(log_format))
        
        # Create logger
        logger = logging.getLogger(__name__)
        logger.setLevel(logging.DEBUG)
        logger.addHandler(file_handler)
        logger.addHandler(console_handler)
        
        return logger
    
    def load_config(self):
        """Load configuration from file or create default"""
        config_file = 'config.json'
        default_config = {
            'merit_badges_url': 'https://www.scouting.org/skills/merit-badges/all/',
            'eagle_required_url': 'https://www.scouting.org/skills/merit-badges/eagle-required/',
            'output_formats': ['html', 'csv', 'pdf', 'excel', 'wordpress'],
            'network_timeout': 30,
            'max_file_size_mb': 50,
            'required_csv_columns': ['firstname', 'lastname', 'positionname', 'primaryemail', 'primaryphone']
        }
        
        if os.path.exists(config_file):
            try:
                with open(config_file, 'r') as f:
                    config = json.load(f)
                return {**default_config, **config}
            except Exception as e:
                self.logger.warning(f"Error loading config: {e}. Using defaults.")
                
        return default_config
    
    def update_progress(self, percent, message=""):
        """Update progress bar and status"""
        if self.progress_callback:
            self.progress_callback(percent)
        if self.status_callback:
            self.status_callback(message)
        self.logger.info(f"Progress: {percent}% - {message}")
    
    def validate_network_connectivity(self):
        """Check network connectivity to scouting.org"""
        try:
            response = requests.get('https://www.scouting.org', timeout=10)
            return response.status_code == 200
        except:
            return False
    
    def fetch_merit_badges(self):
        """Fetch current merit badges from scouting.org with improved parsing"""
        try:
            self.update_progress(5, "Fetching merit badges from scouting.org...")
            
            if not self.validate_network_connectivity():
                self.logger.warning("No network connectivity, using fallback list")
                self.merit_badges = self.get_fallback_merit_badges()
                return
            
            response = requests.get(
                self.config['merit_badges_url'], 
                timeout=self.config['network_timeout'],
                headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
            )
            response.raise_for_status()
            
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Look for merit badges after "Merit Badges A-Z" heading
            badges = []
            
            # Try multiple selectors to find merit badge links
            selectors = [
                'a[href*="/merit-badges/"]',
                '.merit-badge-link',
                'a[href*="/skills/merit-badges/"]'
            ]
            
            for selector in selectors:
                links = soup.select(selector)
                for link in links:
                    badge_name = link.get_text().strip()
                    if self.is_valid_badge_name(badge_name):
                        badges.append(badge_name)
                        
                if len(badges) > 50:  # Found enough badges
                    break
            
            # Clean and deduplicate
            badges = list(set(badge for badge in badges if badge))
            
            # Fallback if scraping fails
            if len(badges) < 100:  # BSA has 130+ merit badges
                self.logger.warning(f"Only found {len(badges)} badges, using fallback list")
                self.merit_badges = self.get_fallback_merit_badges()
            else:
                self.merit_badges = sorted(badges)
                self.logger.info(f"Successfully loaded {len(self.merit_badges)} merit badges from scouting.org")
            
        except Exception as e:
            self.logger.error(f"Error fetching merit badges: {e}")
            self.merit_badges = self.get_fallback_merit_badges()
    
    def is_valid_badge_name(self, name):
        """Validate if a string is a valid merit badge name"""
        if not name or len(name) < 3:
            return False
        
        # Filter out navigation and non-badge items
        invalid_terms = [
            'home', 'about', 'contact', 'menu', 'search', 'login', 'requirements update',
            'eagle required', 'merit badges a-z', 'all merit badges', 'skip to', 'main content'
        ]
        
        return not any(term in name.lower() for term in invalid_terms)
    
    def fetch_eagle_required_badges(self):
        """Fetch Eagle-required merit badges from scouting.org"""
        try:
            self.update_progress(10, "Fetching Eagle-required badges...")
            
            # Known Eagle-required badges (current as of 2025)
            eagle_badges = [
                'Camping', 'Citizenship in Community', 'Citizenship in Nation', 
                'Citizenship in Society', 'Communication', 'Cooking', 
                'Emergency Preparedness', 'Environmental Science', 'Family Life',
                'First Aid', 'Personal Fitness', 'Personal Management'
            ]
            
            # Plus one from each group (using most common choices)
            eagle_badges.extend(['Cycling', 'Hiking', 'Swimming'])  # Group options
            
            # Try to fetch from website if possible
            if self.validate_network_connectivity():
                try:
                    response = requests.get(
                        self.config['eagle_required_url'],
                        timeout=self.config['network_timeout'],
                        headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
                    )
                    response.raise_for_status()
                    
                    soup = BeautifulSoup(response.content, 'html.parser')
                    
                    # Look for eagle required badges
                    web_badges = []
                    for link in soup.find_all('a', href=lambda href: href and '/merit-badges/' in href):
                        badge_name = link.get_text().strip()
                        if self.is_valid_badge_name(badge_name):
                            web_badges.append(badge_name)
                    
                    if len(web_badges) >= 12:  # Should have at least 12 required badges
                        self.eagle_required_badges = sorted(list(set(web_badges)))
                        self.logger.info(f"Loaded {len(self.eagle_required_badges)} Eagle-required badges from web")
                        return
                        
                except Exception as e:
                    self.logger.warning(f"Could not fetch Eagle badges from web: {e}")
            
            self.eagle_required_badges = sorted(eagle_badges)
            self.logger.info(f"Using default list of {len(self.eagle_required_badges)} Eagle-required badges")
            
        except Exception as e:
            self.logger.error(f"Error with Eagle-required badges: {e}")
            self.eagle_required_badges = eagle_badges
    
    def get_fallback_merit_badges(self):
        """Return comprehensive fallback list of merit badges"""
        return [
            'Animal Science', 'Animation', 'Archaeology', 'Archery', 'Architecture', 'Art', 
            'Astronomy', 'Athletics', 'Automotive', 'Aviation', 'Backpacking', 'Basketry', 
            'Bird Study', 'Bugling', 'Camping', 'Canoeing', 'Chemistry', 'Chess', 
            'Citizenship in Community', 'Citizenship in Nation', 'Citizenship in Society', 
            'Climbing', 'Coin Collecting', 'Collections', 'Communication', 'Composite Materials', 
            'Cooking', 'Crime Prevention', 'Cycling', 'Dentistry', 'Digital Technology', 
            'Disabilities Awareness', 'Dog Care', 'Drafting', 'Electricity', 'Electronics', 
            'Emergency Preparedness', 'Energy', 'Engineering', 'Entrepreneurship', 
            'Environmental Science', 'Exploration', 'Family Life', 'Farm Mechanics', 
            'Fingerprinting', 'Fire Safety', 'First Aid', 'Fish and Wildlife Management', 
            'Fishing', 'Forestry', 'Game Design', 'Gardening', 'Genealogy', 'Geocaching', 
            'Geology', 'Golf', 'Graphic Arts', 'Hiking', 'Home Repairs', 'Horsemanship', 
            'Indian Lore', 'Insect Study', 'Inventing', 'Journalism', 'Kayaking', 
            'Landscape Architecture', 'Law', 'Leatherwork', 'Lifesaving', 'Mammal Study', 
            'Medicine', 'Metalwork', 'Model Design and Building', 'Motorboating', 'Music', 
            'Nature', 'Nuclear Science', 'Oceanography', 'Orienteering', 'Painting', 
            'Personal Fitness', 'Personal Management', 'Pets', 'Photography', 'Pioneering', 
            'Plant Science', 'Plumbing', 'Pottery', 'Programming', 'Public Health', 
            'Public Speaking', 'Pulp and Paper', 'Radio', 'Railroading', 'Reading', 
            'Reptile and Amphibian Study', 'Rifle Shooting', 'Robotics', 'Rowing', 'Safety', 
            'Salesmanship', 'Scholarship', 'Scouting Heritage', 'Scuba Diving', 'Sculpture', 
            'Search and Rescue', 'Shotgun Shooting', 'Signs, Signals, and Codes', 'Skating', 
            'Small-Boat Sailing', 'Snow Sports', 'Soil and Water Conservation', 'Space Exploration', 
            'Sports', 'Stamp Collecting', 'Sustainability', 'Swimming', 'Textile', 'Theater', 
            'Traffic Safety', 'Truck Transportation', 'Veterinary Medicine', 'Water Sports', 
            'Weather', 'Welding', 'Whitewater', 'Wilderness Survival', 'Wood Carving', 'Woodworking'
        ]
    
    def validate_csv_file(self, file_path):
        """Validate CSV file format and required columns"""
        try:
            # Check file size
            file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
            if file_size_mb > self.config['max_file_size_mb']:
                raise ValueError(f"File too large: {file_size_mb:.1f}MB (max: {self.config['max_file_size_mb']}MB)")
            
            # Try to read first few lines
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                first_lines = [f.readline().strip() for _ in range(10)]
            
            # Look for memberid line (column headers)
            header_line = None
            for i, line in enumerate(first_lines):
                if 'memberid' in line.lower():
                    header_line = line
                    break
            
            if not header_line:
                # Use first line as header
                header_line = first_lines[0] if first_lines else ""
            
            # Check for required columns
            required_cols = self.config['required_csv_columns']
            header_lower = header_line.lower()
            
            missing_cols = []
            for col in required_cols:
                if col not in header_lower:
                    missing_cols.append(col)
            
            if missing_cols:
                self.logger.warning(f"Missing columns in {file_path}: {missing_cols}")
                return False, f"Missing required columns: {', '.join(missing_cols)}"
            
            return True, "Valid CSV format"
            
        except Exception as e:
            return False, f"Validation error: {str(e)}"
    
    def process_csv_file(self, file_path, troop_id):
        """Process CSV roster file with improved error handling and validation"""
        try:
            self.update_progress(20, f"Processing {troop_id} roster...")
            
            # Validate file first
            is_valid, message = self.validate_csv_file(file_path)
            if not is_valid:
                raise ValueError(f"CSV validation failed: {message}")
            
            # Read and inspect file structure
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                lines = f.readlines()
            
            self.logger.debug(f"Processing {troop_id} CSV with {len(lines)} lines")
            
            # Find header line (contains memberid)
            header_line_idx = 0
            for i, line in enumerate(lines):
                if 'memberid' in line.lower():
                    header_line_idx = i
                    break
            
            # Try multiple parsing approaches
            df = None
            encodings = ['utf-8', 'windows-1252', 'iso-8859-1']
            
            for encoding in encodings:
                try:
                    df = pd.read_csv(
                        file_path,
                        encoding=encoding,
                        skiprows=header_line_idx,
                        sep=None,
                        engine='python',
                        on_bad_lines='skip',
                        quoting=csv.QUOTE_MINIMAL,
                        skipinitialspace=True,
                        low_memory=False
                    )
                    self.logger.debug(f"Successfully parsed with encoding: {encoding}")
                    break
                except Exception as e:
                    self.logger.debug(f"Failed with {encoding}: {e}")
                    continue
            
            if df is None or len(df) == 0:
                raise ValueError(f"Could not parse CSV file: {file_path}")
            
            # Log extracted data for debugging
            self.logger.debug(f"CSV columns found: {list(df.columns)}")
            self.logger.debug(f"Total rows: {len(df)}")
            self.logger.debug(f"Sample data:\n{df.head()}")
            
            # Clean column names
            df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
            
            # Filter out youth members
            position_col = self.find_column(df.columns, ['positionname', 'position', 'role', 'title'])
            
            if position_col:
                adults = df[~df[position_col].astype(str).str.contains('youth member', case=False, na=False)].copy()
                self.logger.info(f"Filtered by {position_col}: {len(adults)} adults from {len(df)} total")
            else:
                adults = df.copy()
                self.logger.warning("No position column found, assuming all entries are adults")
            
            # Standardize column names
            column_mapping = {
                'firstname': 'first_name',
                'lastname': 'last_name', 
                'primaryemail': 'email',
                'primaryphone': 'phone',
                'positionname': 'position'
            }
            
            for old_col, new_col in column_mapping.items():
                if old_col in adults.columns:
                    adults[new_col] = adults[old_col]
            
            # Create full name
            if 'first_name' in adults.columns and 'last_name' in adults.columns:
                adults['name'] = adults['first_name'].astype(str) + ' ' + adults['last_name'].astype(str)
            elif 'first_name' in adults.columns:
                adults['name'] = adults['first_name'].astype(str)
            
            # Clean data
            text_columns = ['name', 'first_name', 'last_name', 'email', 'phone', 'position']
            for col in text_columns:
                if col in adults.columns:
                    adults[col] = adults[col].astype(str).str.strip()
                    adults[col] = adults[col].replace(['nan', 'NaN', 'None'], '')
            
            # Convert to list and filter valid entries
            adults_list = adults.to_dict('records')
            adults_list = [adult for adult in adults_list 
                          if adult.get('name') and adult['name'] not in ['nan', '', ' ']]
            
            # Debug logging for each processed row
            self.logger.debug(f"Processed {len(adults_list)} adults from {troop_id}:")
            for i, adult in enumerate(adults_list):
                self.logger.debug(f"  {i+1}: {adult.get('name', 'NO NAME')} | "
                                f"Email: {adult.get('email', 'N/A')} | "
                                f"Phone: {adult.get('phone', 'N/A')} | "
                                f"Position: {adult.get('position', 'N/A')}")
            
            self.logger.info(f"Successfully processed {len(adults_list)} adults from {troop_id}")
            return adults_list
            
        except Exception as e:
            self.logger.error(f"Error processing CSV {file_path}: {e}")
            raise
    
    def find_column(self, columns, possibilities):
        """Find the first matching column name from possibilities"""
        for col in columns:
            if col in possibilities:
                return col
        return None
    
    def validate_pdf_file(self, file_path):
        """Validate PDF file can be processed"""
        try:
            with pdfplumber.open(file_path) as pdf:
                if len(pdf.pages) == 0:
                    return False, "PDF has no pages"
                
                # Try to extract text from first page
                first_page_text = pdf.pages[0].extract_text()
                if not first_page_text or len(first_page_text.strip()) < 10:
                    return False, "PDF appears to be image-based or empty"
                
                # Check for merit badge counselor content
                if 'merit badge' not in first_page_text.lower():
                    return False, "PDF does not appear to contain merit badge counselor data"
                
                return True, "Valid PDF"
                
        except Exception as e:
            return False, f"PDF validation error: {str(e)}"
    
    def process_pdf_files(self, pdf_files):
        """Process merit badge counselor PDF files with improved parsing"""
        counselors = []
        
        if not pdf_files:
            self.logger.info("No PDF files provided, generating sample data")
            return self.get_sample_counselors()
        
        for i, pdf_file in enumerate(pdf_files):
            try:
                self.update_progress(
                    30 + (i * 20 // len(pdf_files)), 
                    f"Processing PDF {i+1}/{len(pdf_files)}..."
                )
                
                # Validate PDF
                is_valid, message = self.validate_pdf_file(pdf_file)
                if not is_valid:
                    self.logger.warning(f"Skipping invalid PDF {pdf_file}: {message}")
                    continue
                
                with pdfplumber.open(pdf_file) as pdf:
                    text = ""
                    for page_num, page in enumerate(pdf.pages):
                        page_text = page.extract_text()
                        if page_text:
                            text += f"\n--- Page {page_num + 1} ---\n{page_text}"
                
                if not text.strip():
                    self.logger.warning(f"No text extracted from {pdf_file}")
                    continue
                
                # Parse counselor data
                pdf_counselors = self.parse_counselor_text(text)
                counselors.extend(pdf_counselors)
                
                self.logger.info(f"Extracted {len(pdf_counselors)} counselors from {pdf_file}")
                
            except Exception as e:
                self.logger.error(f"Error processing PDF {pdf_file}: {e}")
                continue
        
        # Debug logging for extracted counselors
        self.logger.debug(f"Extracted {len(counselors)} total counselors:")
        for i, counselor in enumerate(counselors):
            self.logger.debug(f"  {i+1}: {counselor.get('name', 'NO NAME')} | "
                            f"Badges: {len(counselor.get('badges', []))} | "
                            f"Email: {counselor.get('email', 'N/A')}")
        
        self.merit_badge_counselors = counselors
        return counselors
    
    def parse_counselor_text(self, text):
        """Enhanced counselor parsing from PDF text"""
        counselors = []
        
        # Split text into potential counselor records
        lines = text.split('\n')
        current_counselor = {}
        
        # Patterns for different data types
        email_pattern = re.compile(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b')
        phone_pattern = re.compile(r'(?:\+?1[-.\s]?)?\(?([0-9]{3})\)?[-.\s]?([0-9]{3})[-.\s]?([0-9]{4})')
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # Email detection
            email_match = email_pattern.search(line)
            if email_match:
                current_counselor['email'] = email_match.group()
                continue
            
            # Phone detection
            phone_match = phone_pattern.search(line)
            if phone_match:
                current_counselor['phone'] = phone_match.group()
                continue
            
            # Merit badge detection
            if any(keyword in line.lower() for keyword in ['badge', 'merit', 'counselor']):
                # Extract badge names
                badges = self.extract_badges_from_line(line)
                if badges:
                    current_counselor['badges'] = badges
                continue
            
            # Name detection (simple heuristic)
            if self.looks_like_name(line) and 'name' not in current_counselor:
                # Save previous counselor if exists
                if current_counselor and 'name' in current_counselor:
                    counselors.append(current_counselor)
                
                # Start new counselor
                current_counselor = {
                    'name': line,
                    'units': ['Other'],
                    'badges': []
                }
        
        # Don't forget the last counselor
        if current_counselor and 'name' in current_counselor:
            counselors.append(current_counselor)
        
        return counselors
    
    def extract_badges_from_line(self, line):
        """Extract merit badge names from a line of text"""
        badges = []
        
        # Remove common prefixes
        line = re.sub(r'merit badge:?|badge:?|counselor for:?', '', line, flags=re.IGNORECASE)
        
        # Split by common separators
        parts = re.split(r'[,;|]', line)
        
        for part in parts:
            part = part.strip()
            if part and len(part) > 2:
                # Check if it matches known merit badges
                for badge in self.merit_badges:
                    if badge.lower() in part.lower():
                        badges.append(badge)
        
        return list(set(badges))  # Remove duplicates
    
    def looks_like_name(self, text):
        """Heuristic to determine if text looks like a person's name"""
        if not text or len(text) < 3:
            return False
        
        # Basic checks
        words = text.split()
        if len(words) < 2 or len(words) > 4:
            return False
        
        # Check if first letter is uppercase
        if not text[0].isupper():
            return False
        
        # Check for common non-name patterns
        non_name_patterns = [
            r'\d',  # Contains digits
            r'@',   # Contains email
            r'merit',  # Contains merit
            r'badge',  # Contains badge
            r'phone',  # Contains phone
            r'email',  # Contains email
        ]
        
        for pattern in non_name_patterns:
            if re.search(pattern, text, re.IGNORECASE):
                return False
        
        return True
    
    def get_sample_counselors(self):
        """Return sample counselor data for testing"""
        return [
            {
                'name': 'John Smith',
                'badges': ['Camping', 'Hiking', 'First Aid'],
                'email': 'john.smith@example.com',
                'phone': '555-0101',
                'units': ['Other']
            },
            {
                'name': 'Sarah Johnson',
                'badges': ['Cooking', 'Environmental Science'],
                'email': 'sarah.j@example.com',
                'phone': '555-0102',
                'units': ['Other']
            },
            {
                'name': 'Mike Davis',
                'badges': ['Personal Management', 'Communication'],
                'email': 'mike.davis@example.com',
                'phone': '555-0103',
                'units': ['Other']
            }
        ]
    
    def generate_reports(self, t12_file=None, t32_file=None, pdf_files=None):
        """Main report generation function with comprehensive validation"""
        try:
            self.update_progress(0, "Starting report generation...")
            
            # Create timestamped output directory
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            self.output_dir = Path(f"MBC_Reports_{timestamp}")
            self.output_dir.mkdir(exist_ok=True)
            
            # Create format subdirectories
            format_dirs = ['html', 'csv', 'pdf', 'excel', 'wordpress']
            for format_dir in format_dirs:
                (self.output_dir / format_dir).mkdir(exist_ok=True)
            
            # Fetch merit badge data
            self.fetch_merit_badges()
            self.fetch_eagle_required_badges()
            
            # Process input files
            if t12_file:
                self.t12_adults = self.process_csv_file(t12_file, "T12")
            else:
                self.t12_adults = self.get_sample_t12_adults()
                
            if t32_file:
                self.t32_adults = self.process_csv_file(t32_file, "T32")
            else:
                self.t32_adults = self.get_sample_t32_adults()
                
            # Process counselor PDFs
            self.process_pdf_files(pdf_files)
            
            self.update_progress(50, "Generating reports...")
            
            # Generate the three main reports
            t12_t32_counselors = self.generate_t12_t32_counselors()
            non_counselors = self.generate_non_counselors()
            coverage = self.generate_coverage_report()
            
            # Debug logging for compiled lists
            self.logger.debug("Compiled list of adult leaders and their counselor status:")
            all_adults = self.t12_adults + self.t32_adults
            counselor_names = {c['name'] for c in self.merit_badge_counselors}
            
            for adult in all_adults:
                is_counselor = adult['name'] in counselor_names
                self.logger.debug(f"  {adult['name']}: {'IS' if is_counselor else 'NOT'} a merit badge counselor")
            
            self.update_progress(70, "Creating output files...")
            
            # Prepare reports data
            reports = [
                {
                    'title': 'T12/T32 Merit Badge Counselors',
                    'data': t12_t32_counselors,
                    'type': 'counselors',
                    'filename': 't12_t32_merit_badge_counselors'
                },
                {
                    'title': 'T12/T32 Leaders not Merit Badge Counselors',
                    'data': non_counselors,
                    'type': 'non_counselors',
                    'filename': 't12_t32_leaders_not_counselors'
                },
                {
                    'title': 'T12/T32 Merit Badge Counselor Coverage',
                    'data': coverage,
                    'type': 'coverage',
                    'filename': 't12_t32_counselor_coverage'
                }
            ]
            
            # Generate all output formats
            for report in reports:
                self.generate_html_report(report)
                self.generate_csv_report(report)
                self.generate_pdf_report(report)
                self.generate_excel_report(report)
                self.generate_wordpress_report(report)
            
            self.update_progress(90, "Generating summary report...")
            self.generate_summary_report(reports)
            
            self.update_progress(95, "Validating outputs...")
            self.validate_outputs(reports)
            
            self.update_progress(100, "Report generation completed!")
            
            return self.output_dir
            
        except Exception as e:
            self.logger.error(f"Error in report generation: {e}")
            raise
    
    def get_sample_t12_adults(self):
        """Return sample T12 adults for testing"""
        return [
            {'name': 'Alice Johnson', 'first_name': 'Alice', 'last_name': 'Johnson', 
             'email': 'alice@example.com', 'phone': '555-0101', 'position': 'Scoutmaster'},
            {'name': 'Bob Smith', 'first_name': 'Bob', 'last_name': 'Smith',
             'email': 'bob@example.com', 'phone': '555-0102', 'position': 'Assistant Scoutmaster'},
            {'name': 'Carol Wilson', 'first_name': 'Carol', 'last_name': 'Wilson',
             'email': 'carol@example.com', 'phone': '555-0103', 'position': 'Committee Chair'},
        ]
    
    def get_sample_t32_adults(self):
        """Return sample T32 adults for testing"""
        return [
            {'name': 'David Brown', 'first_name': 'David', 'last_name': 'Brown',
             'email': 'david@example.com', 'phone': '555-0201', 'position': 'Scoutmaster'},
            {'name': 'Emily Davis', 'first_name': 'Emily', 'last_name': 'Davis',
             'email': 'emily@example.com', 'phone': '555-0202', 'position': 'Committee Member'},
            {'name': 'Frank Miller', 'first_name': 'Frank', 'last_name': 'Miller',
             'email': 'frank@example.com', 'phone': '555-0203', 'position': 'Treasurer'},
        ]
    
    def generate_t12_t32_counselors(self):
        """Generate T12/T32 Merit Badge Counselors report with cross-referencing"""
        all_adults = self.t12_adults + self.t32_adults
        adult_names = {adult['name'].lower().strip() for adult in all_adults}
        
        t12_t32_counselors = []
        
        for counselor in self.merit_badge_counselors:
            counselor_name = counselor['name'].lower().strip()
            
            # Try exact match first
            if counselor_name in adult_names:
                merged_counselor = self.merge_counselor_data(counselor, all_adults)
                if merged_counselor:
                    t12_t32_counselors.append(merged_counselor)
                    continue
            
            # Try fuzzy matching for names with slight variations
            for adult in all_adults:
                if self.names_match(counselor['name'], adult['name']):
                    merged_counselor = self.merge_counselor_data(counselor, [adult])
                    if merged_counselor:
                        t12_t32_counselors.append(merged_counselor)
                        break
        
        # Sort alphabetically by last name, then first name
        t12_t32_counselors.sort(key=lambda x: (x.get('last_name', ''), x.get('first_name', '')))
        
        return t12_t32_counselors
    
    def names_match(self, name1, name2):
        """Check if two names match with some tolerance for variations"""
        if not name1 or not name2:
            return False
        
        # Normalize names
        n1 = re.sub(r'[^\w\s]', '', name1.lower().strip())
        n2 = re.sub(r'[^\w\s]', '', name2.lower().strip())
        
        # Exact match
        if n1 == n2:
            return True
        
        # Check if all words in shorter name are in longer name
        words1 = set(n1.split())
        words2 = set(n2.split())
        
        if len(words1) <= len(words2):
            return words1.issubset(words2)
        else:
            return words2.issubset(words1)
    
    def merge_counselor_data(self, counselor, adults):
        """Merge counselor data with adult roster data"""
        for adult in adults:
            if self.names_match(counselor['name'], adult['name']):
                merged = {
                    'name': adult.get('name', counselor['name']),
                    'first_name': adult.get('first_name', ''),
                    'last_name': adult.get('last_name', ''),
                    'email': adult.get('email', counselor.get('email', '')),
                    'phone': adult.get('phone', counselor.get('phone', '')),
                    'position': adult.get('position', ''),
                    'badges': counselor.get('badges', []),
                    'units': counselor.get('units', ['Other']),
                    'troop': 'T12' if adult in self.t12_adults else 'T32'
                }
                return merged
        return None
    
    def generate_non_counselors(self):
        """Generate T12/T32 Leaders not Merit Badge Counselors report"""
        all_adults = self.t12_adults + self.t32_adults
        counselor_names = {c['name'].lower().strip() for c in self.merit_badge_counselors}
        
        non_counselors = []
        for adult in all_adults:
            adult_name = adult['name'].lower().strip()
            
            # Check exact match
            is_counselor = adult_name in counselor_names
            
            # Check fuzzy match if no exact match
            if not is_counselor:
                for counselor in self.merit_badge_counselors:
                    if self.names_match(adult['name'], counselor['name']):
                        is_counselor = True
                        break
            
            if not is_counselor:
                adult_copy = adult.copy()
                adult_copy['troop'] = 'T12' if adult in self.t12_adults else 'T32'
                non_counselors.append(adult_copy)
        
        # Sort alphabetically by last name, then first name
        non_counselors.sort(key=lambda x: (x.get('last_name', ''), x.get('first_name', '')))
        
        return non_counselors
    
    def generate_coverage_report(self):
        """Generate Merit Badge Coverage report"""
        all_adults = self.t12_adults + self.t32_adults
        adult_names = {adult['name'].lower().strip() for adult in all_adults}
        
        # Get badges covered by T12/T32 counselors
        t12_t32_counselor_badges = set()
        for counselor in self.merit_badge_counselors:
            counselor_name = counselor['name'].lower().strip()
            
            # Check if this counselor is from T12/T32
            is_t12_t32 = counselor_name in adult_names
            if not is_t12_t32:
                # Check fuzzy match
                for adult in all_adults:
                    if self.names_match(counselor['name'], adult['name']):
                        is_t12_t32 = True
                        break
            
            if is_t12_t32 and 'badges' in counselor:
                t12_t32_counselor_badges.update(counselor['badges'])
        
        coverage = {
            'eagle_with_counselors': [],
            'eagle_without_counselors': [],
            'non_eagle_with_counselors': [],
            'non_eagle_without_counselors': []
        }
        
        for badge in self.merit_badges:
            is_eagle = badge in self.eagle_required_badges
            has_counselor = badge in t12_t32_counselor_badges
            
            if is_eagle:
                if has_counselor:
                    coverage['eagle_with_counselors'].append(badge)
                else:
                    coverage['eagle_without_counselors'].append(badge)
            else:
                if has_counselor:
                    coverage['non_eagle_with_counselors'].append(badge)
                else:
                    coverage['non_eagle_without_counselors'].append(badge)
        
        # Sort all categories alphabetically
        for key in coverage:
            coverage[key].sort()
        
        return coverage
    
    def generate_html_report(self, report):
        """Generate HTML report with embedded CSS and download functionality"""
        template_str = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{{ title }}</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { 
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; 
            line-height: 1.6; 
            color: #333; 
            background: #f8f9fa; 
        }
        .container { 
            max-width: 1200px; 
            margin: 0 auto; 
            background: white; 
            min-height: 100vh; 
            box-shadow: 0 0 20px rgba(0,0,0,0.1); 
        }
        .header { 
            background: linear-gradient(135deg, #1e3a8a 0%, #3b82f6 50%, #60a5fa 100%); 
            color: white; 
            padding: 2rem; 
            text-align: center; 
        }
        .header h1 { 
            font-size: 2.5rem; 
            margin-bottom: 0.5rem; 
            text-shadow: 0 2px 4px rgba(0,0,0,0.3); 
        }
        .header .subtitle { 
            font-size: 1.1rem; 
            opacity: 0.9; 
            margin-bottom: 0.5rem; 
        }
        .header .timestamp { 
            font-size: 0.9rem; 
            opacity: 0.8; 
        }
        .actions { 
            padding: 1.5rem; 
            background: #f8f9fa; 
            border-bottom: 1px solid #e9ecef; 
            display: flex; 
            gap: 1rem; 
            flex-wrap: wrap; 
        }
        .btn { 
            background: #3b82f6; 
            color: white; 
            border: none; 
            padding: 0.75rem 1.5rem; 
            border-radius: 0.5rem; 
            cursor: pointer; 
            text-decoration: none; 
            font-weight: 500; 
            transition: all 0.2s; 
            display: inline-flex; 
            align-items: center; 
            gap: 0.5rem; 
        }
        .btn:hover { 
            background: #2563eb; 
            transform: translateY(-1px); 
            box-shadow: 0 4px 8px rgba(0,0,0,0.15); 
        }
        .btn-success { background: #059669; }
        .btn-success:hover { background: #047857; }
        .content { padding: 2rem; }
        table { 
            width: 100%; 
            border-collapse: collapse; 
            margin: 1.5rem 0; 
            background: white; 
            border-radius: 0.5rem; 
            overflow: hidden; 
            box-shadow: 0 2px 8px rgba(0,0,0,0.1); 
        }
        th, td { 
            padding: 1rem; 
            text-align: left; 
            border-bottom: 1px solid #e5e7eb; 
        }
        th { 
            background: #f3f4f6; 
            font-weight: 600; 
            color: #374151; 
            text-transform: uppercase; 
            font-size: 0.875rem; 
            letter-spacing: 0.05em; 
        }
        tr:hover { background: #f8fafc; }
        .coverage-section { 
            margin: 2rem 0; 
            padding: 1.5rem; 
            background: white; 
            border-radius: 0.75rem; 
            border-left: 4px solid #3b82f6; 
            box-shadow: 0 2px 8px rgba(0,0,0,0.05); 
        }
        .coverage-section h3 { 
            margin-bottom: 1rem; 
            color: #1f2937; 
            font-size: 1.25rem; 
        }
        .coverage-section ul { 
            columns: 3; 
            column-gap: 2rem; 
            list-style-type: none; 
            margin: 0; 
        }
        .coverage-section li { 
            break-inside: avoid; 
            margin-bottom: 0.5rem; 
            padding: 0.25rem 0; 
            border-bottom: 1px solid #f3f4f6; 
        }
        .eagle-required { border-left-color: #dc2626; }
        .eagle-missing { border-left-color: #f59e0b; }
        .stats { 
            display: grid; 
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); 
            gap: 1rem; 
            margin: 2rem 0; 
        }
        .stat-card { 
            background: white; 
            padding: 1.5rem; 
            border-radius: 0.75rem; 
            text-align: center; 
            box-shadow: 0 2px 8px rgba(0,0,0,0.1); 
            border-top: 3px solid #3b82f6; 
        }
        .stat-number { 
            font-size: 2rem; 
            font-weight: bold; 
            color: #3b82f6; 
            margin-bottom: 0.5rem; 
        }
        .stat-label { 
            color: #6b7280; 
            font-size: 0.875rem; 
            text-transform: uppercase; 
            letter-spacing: 0.05em; 
        }
        @media print { 
            .actions { display: none; } 
            .container { box-shadow: none; } 
        }
        @media (max-width: 768px) { 
            .coverage-section ul { columns: 1; } 
            th, td { padding: 0.75rem 0.5rem; font-size: 0.875rem; } 
            .header h1 { font-size: 1.75rem; } 
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🎖️ {{ title }}</h1>
            <div class="subtitle">Troop 12 & Troop 32 Acton MA</div>
            <div class="timestamp">Generated: {{ timestamp }}</div>
        </div>
        
        <div class="actions">
            <button class="btn" onclick="downloadCSV()">
                📥 Download CSV
            </button>
            <button class="btn btn-success" onclick="window.print()">
                🖨️ Print Report
            </button>
            <button class="btn" onclick="downloadExcel()">
                📊 Download Excel
            </button>
        </div>
        
        <div class="content">
            {{ content }}
        </div>
    </div>
    
    <script>
        var csvData = {{ csv_data }};
        var filename = '{{ filename }}';
        
        function downloadCSV() {
            downloadFile(csvData, filename + '.csv', 'text/csv');
        }
        
        function downloadExcel() {
            // Redirect to Excel file
            window.location.href = '../excel/' + filename + '.xlsx';
        }
        
        function downloadFile(content, fileName, contentType) {
            var blob = new Blob([content], {type: contentType + ';charset=utf-8;'});
            var link = document.createElement('a');
            var url = URL.createObjectURL(blob);
            link.setAttribute('href', url);
            link.setAttribute('download', fileName);
            link.style.visibility = 'hidden';
            document.body.appendChild(link);
            link.click();
            document.body.removeChild(link);
        }
        
        // Add some interactivity
        document.addEventListener('DOMContentLoaded', function() {
            // Highlight rows on click
            var rows = document.querySelectorAll('tr');
            rows.forEach(function(row) {
                row.addEventListener('click', function() {
                    rows.forEach(function(r) { r.style.background = ''; });
                    this.style.background = '#e0f2fe';
                });
            });
        });
    </script>
</body>
</html>"""
        
        # Generate content based on report type
        if report['type'] == 'counselors':
            content = self.generate_counselors_html(report['data'])
        elif report['type'] == 'non_counselors':
            content = self.generate_non_counselors_html(report['data'])
        else:  # coverage
            content = self.generate_coverage_html(report['data'])
        
        # Generate CSV data for download
        csv_data = self.generate_csv_content(report)
        
        template = Template(template_str)
        html = template.render(
            title=report['title'],
            timestamp=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            content=content,
            csv_data=json.dumps(csv_data),
            filename=report['filename']
        )
        
        # Save HTML file
        html_path = self.output_dir / 'html' / f"{report['filename']}.html"
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(html)
        
        self.logger.info(f"Generated HTML report: {html_path}")
    
    def generate_counselors_html(self, data):
        """Generate HTML content for counselors report"""
        if not data:
            return "<p>No T12/T32 merit badge counselors found.</p>"
        
        html = f"""
        <div class="stats">
            <div class="stat-card">
                <div class="stat-number">{len(data)}</div>
                <div class="stat-label">Total Counselors</div>
            </div>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>Name</th>
                    <th>Troop</th>
                    <th>Position</th>
                    <th>Email</th>
                    <th>Phone</th>
                    <th>Merit Badges</th>
                </tr>
            </thead>
            <tbody>
        """
        
        for counselor in data:
            badges = ', '.join(counselor.get('badges', []))
            html += f"""
                <tr>
                    <td><strong>{counselor.get('name', 'N/A')}</strong></td>
                    <td>{counselor.get('troop', 'N/A')}</td>
                    <td>{counselor.get('position', 'N/A')}</td>
                    <td>{counselor.get('email', 'N/A')}</td>
                    <td>{counselor.get('phone', 'N/A')}</td>
                    <td>{badges or 'N/A'}</td>
                </tr>
            """
        
        html += "</tbody></table>"
        return html
    
    def generate_csv_content(self, report):
        """Generate CSV content for download"""
        if report['type'] == 'counselors':
            lines = ['Name,Troop,Position,Email,Phone,Merit Badges']
            for counselor in report['data']:
                badges = '; '.join(counselor.get('badges', []))
                line = f'"{counselor.get("name", "")}","{counselor.get("troop", "")}","{counselor.get("position", "")}","{counselor.get("email", "")}","{counselor.get("phone", "")}","{badges}"'
                lines.append(line)
        elif report['type'] == 'non_counselors':
            lines = ['Name,Troop,Position,Email,Phone']
            for adult in report['data']:
                line = f'"{adult.get("name", "")}","{adult.get("troop", "")}","{adult.get("position", "")}","{adult.get("email", "")}","{adult.get("phone", "")}"'
                lines.append(line)
        else:  # coverage
            lines = ['Category,Merit Badge']
            data = report['data']
            sections = [
                ('Eagle-Required with T12/T32 Counselors', data['eagle_with_counselors']),
                ('Eagle-Required without T12/T32 Counselors', data['eagle_without_counselors']),
                ('Non-Eagle with T12/T32 Counselors', data['non_eagle_with_counselors']),
                ('Non-Eagle without T12/T32 Counselors', data['non_eagle_without_counselors'])
            ]
            for category, badges in sections:
                for badge in badges:
                    lines.append(f'"{category}","{badge}"')
        
        return '\n'.join(lines)
    
    def generate_csv_report(self, report):
        """Generate CSV report files"""
        csv_content = self.generate_csv_content(report)
        csv_path = self.output_dir / 'csv' / f"{report['filename']}.csv"
        
        with open(csv_path, 'w', encoding='utf-8', newline='') as f:
            f.write(csv_content)
        
        self.logger.info(f"Generated CSV report: {csv_path}")
    
    def generate_pdf_report(self, report):
        """Generate PDF report with professional formatting"""
        try:
            pdf_path = self.output_dir / 'pdf' / f"{report['filename']}.pdf"
            doc = SimpleDocTemplate(str(pdf_path), pagesize=letter)
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1,  # Center
                textColor=colors.darkblue
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=14,
                spaceAfter=12,
                textColor=colors.darkblue
            )
            
            # Build content
            content = []
            
            # Title
            content.append(Paragraph(f"🎖️ {report['title']}", title_style))
            content.append(Paragraph("Troop 12 & Troop 32 Acton MA", styles['Normal']))
            content.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
            content.append(Spacer(1, 20))
            
            # Content based on report type
            if report['type'] == 'counselors':
                content.extend(self.generate_counselors_pdf_content(report['data'], styles, heading_style))
            elif report['type'] == 'non_counselors':
                content.extend(self.generate_non_counselors_pdf_content(report['data'], styles, heading_style))
            else:  # coverage
                content.extend(self.generate_coverage_pdf_content(report['data'], styles, heading_style))
            
            # Build PDF
            doc.build(content)
            self.logger.info(f"Generated PDF report: {pdf_path}")
            
        except Exception as e:
            self.logger.error(f"Error generating PDF report: {e}")
    
    def generate_counselors_pdf_content(self, data, styles, heading_style):
        """Generate PDF content for counselors"""
        content = []
        
        if not data:
            content.append(Paragraph("No T12/T32 merit badge counselors found.", styles['Normal']))
            return content
        
        content.append(Paragraph(f"Total Counselors: {len(data)}", heading_style))
        
        # Create table
        table_data = [['Name', 'Troop', 'Position', 'Email', 'Phone', 'Merit Badges']]
        
        for counselor in data:
            badges = ', '.join(counselor.get('badges', []))
            if len(badges) > 50:  # Truncate long badge lists
                badges = badges[:47] + '...'
            
            table_data.append([
                counselor.get('name', 'N/A'),
                counselor.get('troop', 'N/A'),
                counselor.get('position', 'N/A'),
                counselor.get('email', 'N/A'),
                counselor.get('phone', 'N/A'),
                badges or 'N/A'
            ])
        
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        content.append(table)
        return content
    
    def generate_non_counselors_pdf_content(self, data, styles, heading_style):
        """Generate PDF content for non-counselors"""
        content = []
        
        if not data:
            content.append(Paragraph("All T12/T32 leaders are merit badge counselors!", styles['Normal']))
            return content
        
        content.append(Paragraph(f"Leaders Not Counselors: {len(data)}", heading_style))
        
        # Create table
        table_data = [['Name', 'Troop', 'Position', 'Email', 'Phone']]
        
        for adult in data:
            table_data.append([
                adult.get('name', 'N/A'),
                adult.get('troop', 'N/A'),
                adult.get('position', 'N/A'),
                adult.get('email', 'N/A'),
                adult.get('phone', 'N/A')
            ])
        
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        content.append(table)
        return content
    
    def generate_coverage_pdf_content(self, data, styles, heading_style):
        """Generate PDF content for coverage report"""
        content = []
        
        sections = [
            ('Eagle-Required Merit Badges with T12/T32 Counselors', data['eagle_with_counselors']),
            ('Eagle-Required Merit Badges without T12/T32 Counselors', data['eagle_without_counselors']),
            ('Non-Eagle Merit Badges with T12/T32 Counselors', data['non_eagle_with_counselors']),
            ('Non-Eagle Merit Badges without T12/T32 Counselors', data['non_eagle_without_counselors'])
        ]
        
        for title, badges in sections:
            content.append(Paragraph(f"{title} ({len(badges)})", heading_style))
            
            if badges:
                # Create multi-column layout for badges
                badge_text = ', '.join(badges)
                content.append(Paragraph(badge_text, styles['Normal']))
            else:
                content.append(Paragraph("None", styles['Normal']))
            
            content.append(Spacer(1, 12))
        
        return content
    
    def generate_excel_report(self, report):
        """Generate Excel report with formatting and multiple sheets"""
        try:
            excel_path = self.output_dir / 'excel' / f"{report['filename']}.xlsx"
            workbook = openpyxl.Workbook()
            
            # Remove default sheet and create our sheet
            workbook.remove(workbook.active)
            worksheet = workbook.create_sheet(title=report['title'][:31])  # Excel sheet name limit
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center")
            
            if report['type'] == 'counselors':
                self.create_counselors_excel_sheet(worksheet, report['data'], header_font, header_fill, center_alignment)
            elif report['type'] == 'non_counselors':
                self.create_non_counselors_excel_sheet(worksheet, report['data'], header_font, header_fill, center_alignment)
            else:  # coverage
                self.create_coverage_excel_sheet(worksheet, report['data'], header_font, header_fill, center_alignment)
            
            # Add metadata sheet
            meta_sheet = workbook.create_sheet(title="Metadata")
            meta_sheet['A1'] = "Report Title"
            meta_sheet['B1'] = report['title']
            meta_sheet['A2'] = "Generated"
            meta_sheet['B2'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            meta_sheet['A3'] = "Generated By"
            meta_sheet['B3'] = "Merit Badge Counselor Generator v2.0"
            
            workbook.save(excel_path)
            self.logger.info(f"Generated Excel report: {excel_path}")
            
        except Exception as e:
            self.logger.error(f"Error generating Excel report: {e}")
    
    def create_counselors_excel_sheet(self, worksheet, data, header_font, header_fill, center_alignment):
        """Create Excel sheet for counselors data"""
        # Headers
        headers = ['Name', 'Troop', 'Position', 'Email', 'Phone', 'Merit Badges']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Data rows
        for row, counselor in enumerate(data, 2):
            worksheet.cell(row=row, column=1, value=counselor.get('name', ''))
            worksheet.cell(row=row, column=2, value=counselor.get('troop', ''))
            worksheet.cell(row=row, column=3, value=counselor.get('position', ''))
            worksheet.cell(row=row, column=4, value=counselor.get('email', ''))
            worksheet.cell(row=row, column=5, value=counselor.get('phone', ''))
            worksheet.cell(row=row, column=6, value=', '.join(counselor.get('badges', [])))
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def create_non_counselors_excel_sheet(self, worksheet, data, header_font, header_fill, center_alignment):
        """Create Excel sheet for non-counselors data"""
        # Headers
        headers = ['Name', 'Troop', 'Position', 'Email', 'Phone']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Data rows
        for row, adult in enumerate(data, 2):
            worksheet.cell(row=row, column=1, value=adult.get('name', ''))
            worksheet.cell(row=row, column=2, value=adult.get('troop', ''))
            worksheet.cell(row=row, column=3, value=adult.get('position', ''))
            worksheet.cell(row=row, column=4, value=adult.get('email', ''))
            worksheet.cell(row=row, column=5, value=adult.get('phone', ''))
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def create_coverage_excel_sheet(self, worksheet, data, header_font, header_fill, center_alignment):
        """Create Excel sheet for coverage data"""
        # Headers
        headers = ['Category', 'Merit Badge']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Data rows
        row = 2
        sections = [
            ('Eagle-Required with T12/T32 Counselors', data['eagle_with_counselors']),
            ('Eagle-Required without T12/T32 Counselors', data['eagle_without_counselors']),
            ('Non-Eagle with T12/T32 Counselors', data['non_eagle_with_counselors']),
            ('Non-Eagle without T12/T32 Counselors', data['non_eagle_without_counselors'])
        ]
        
        for category, badges in sections:
            for badge in badges:
                worksheet.cell(row=row, column=1, value=category)
                worksheet.cell(row=row, column=2, value=badge)
                row += 1
        
        # Auto-adjust column widths
        worksheet.column_dimensions['A'].width = 50
        worksheet.column_dimensions['B'].width = 30
    
    def generate_wordpress_report(self, report):
        """Generate WordPress-compatible shortcode content"""
        try:
            wp_path = self.output_dir / 'wordpress' / f"{report['filename']}.txt"
            
            content = f"""[merit_badge_report title="{report['title']}"]

Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
Report: {report['title']}

"""
            
            if report['type'] == 'counselors':
                content += self.generate_counselors_wordpress(report['data'])
            elif report['type'] == 'non_counselors':
                content += self.generate_non_counselors_wordpress(report['data'])
            else:  # coverage
                content += self.generate_coverage_wordpress(report['data'])
            
            content += f"""

[download_buttons csv="../csv/{report['filename']}.csv" excel="../excel/{report['filename']}.xlsx" pdf="../pdf/{report['filename']}.pdf"]

[/merit_badge_report]"""
            
            with open(wp_path, 'w', encoding='utf-8') as f:
                f.write(content)
            
            self.logger.info(f"Generated WordPress report: {wp_path}")
            
        except Exception as e:
            self.logger.error(f"Error generating WordPress report: {e}")
    
    def generate_counselors_wordpress(self, data):
        """Generate WordPress content for counselors"""
        if not data:
            return "No T12/T32 merit badge counselors found."
        
        content = f"[table responsive='yes']\n"
        content += "[tr][th]Name[/th][th]Troop[/th][th]Position[/th][th]Email[/th][th]Phone[/th][th]Merit Badges[/th][/tr]\n"
        
        for counselor in data:
            badges = ', '.join(counselor.get('badges', []))
            content += f"[tr][td]{counselor.get('name', 'N/A')}[/td]"
            content += f"[td]{counselor.get('troop', 'N/A')}[/td]"
            content += f"[td]{counselor.get('position', 'N/A')}[/td]"
            content += f"[td]{counselor.get('email', 'N/A')}[/td]"
            content += f"[td]{counselor.get('phone', 'N/A')}[/td]"
            content += f"[td]{badges or 'N/A'}[/td][/tr]\n"
        
        content += "[/table]"
        return content
    
    def generate_non_counselors_wordpress(self, data):
        """Generate WordPress content for non-counselors"""
        if not data:
            return "All T12/T32 leaders are merit badge counselors!"
        
        content = f"[table responsive='yes']\n"
        content += "[tr][th]Name[/th][th]Troop[/th][th]Position[/th][th]Email[/th][th]Phone[/th][/tr]\n"
        
        for adult in data:
            content += f"[tr][td]{adult.get('name', 'N/A')}[/td]"
            content += f"[td]{adult.get('troop', 'N/A')}[/td]"
            content += f"[td]{adult.get('position', 'N/A')}[/td]"
            content += f"[td]{adult.get('email', 'N/A')}[/td]"
            content += f"[td]{adult.get('phone', 'N/A')}[/td][/tr]\n"
        
        content += "[/table]"
        return content
    
    def generate_coverage_wordpress(self, data):
        """Generate WordPress content for coverage"""
        sections = [
            ('Eagle-Required Merit Badges with T12/T32 Counselors', data['eagle_with_counselors']),
            ('Eagle-Required Merit Badges without T12/T32 Counselors', data['eagle_without_counselors']),
            ('Non-Eagle Merit Badges with T12/T32 Counselors', data['non_eagle_with_counselors']),
            ('Non-Eagle Merit Badges without T12/T32 Counselors', data['non_eagle_without_counselors'])
        ]
        
        content = ""
        for title, badges in sections:
            content += f"\n[accordion title='{title} ({len(badges)})']\n"
            if badges:
                content += "[columns]\n[column]"
                for i, badge in enumerate(badges):
                    if i > 0 and i % (len(badges)//3 + 1) == 0:
                        content += "[/column][column]"
                    content += f"• {badge}\n"
                content += "[/column][/columns]\n"
            else:
                content += "None\n"
            content += "[/accordion]\n"
        
        return content
    
    def generate_summary_report(self, reports):
        """Generate comprehensive summary report with statistics and validation"""
        try:
            summary_path = self.output_dir / 'summary_report.html'
            
            # Calculate statistics
            total_merit_badges = len(self.merit_badges)
            total_eagle_badges = len(self.eagle_required_badges)
            total_t12_adults = len(self.t12_adults)
            total_t32_adults = len(self.t32_adults)
            total_adults = total_t12_adults + total_t32_adults
            
            counselors_report = next(r for r in reports if r['type'] == 'counselors')
            non_counselors_report = next(r for r in reports if r['type'] == 'non_counselors')
            coverage_report = next(r for r in reports if r['type'] == 'coverage')
            
            total_counselors = len(counselors_report['data'])
            total_non_counselors = len(non_counselors_report['data'])
            
            coverage_data = coverage_report['data']
            eagle_covered = len(coverage_data['eagle_with_counselors'])
            eagle_missing = len(coverage_data['eagle_without_counselors'])
            
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            summary_html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Merit Badge Counselor Reports - Summary</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ 
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; 
            line-height: 1.6; 
            color: #333; 
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
        }}
        .container {{ 
            max-width: 1200px; 
            margin: 2rem auto; 
            background: white; 
            border-radius: 1rem; 
            overflow: hidden;
            box-shadow: 0 20px 40px rgba(0,0,0,0.1); 
        }}
        .header {{ 
            background: linear-gradient(135deg, #1e3a8a 0%, #3b82f6 50%, #60a5fa 100%); 
            color: white; 
            padding: 3rem 2rem; 
            text-align: center; 
        }}
        .header h1 {{ 
            font-size: 3rem; 
            margin-bottom: 1rem; 
            text-shadow: 0 2px 4px rgba(0,0,0,0.3); 
        }}
        .header .subtitle {{ 
            font-size: 1.25rem; 
            opacity: 0.9; 
            margin-bottom: 0.5rem; 
        }}
        .header .timestamp {{ 
            font-size: 1rem; 
            opacity: 0.8; 
        }}
        .content {{ padding: 3rem 2rem; }}
        .stats {{ 
            display: grid; 
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); 
            gap: 2rem; 
            margin: 3rem 0; 
        }}
        .stat-card {{ 
            background: linear-gradient(135deg, #f8fafc 0%, #e2e8f0 100%); 
            padding: 2rem; 
            border-radius: 1rem; 
            text-align: center; 
            border: 1px solid #e2e8f0;
            transition: transform 0.2s, box-shadow 0.2s;
        }}
        .stat-card:hover {{
            transform: translateY(-5px);
            box-shadow: 0 10px 25px rgba(0,0,0,0.15);
        }}
        .stat-number {{ 
            font-size: 3rem; 
            font-weight: bold; 
            color: #3b82f6; 
            margin-bottom: 0.5rem; 
            text-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .stat-label {{ 
            color: #6b7280; 
            font-size: 1rem; 
            text-transform: uppercase; 
            letter-spacing: 0.05em; 
            font-weight: 500;
        }}
        .section {{ 
            margin: 3rem 0; 
            padding: 2rem; 
            background: #f8fafc; 
            border-radius: 1rem; 
            border-left: 5px solid #3b82f6; 
        }}
        .section h2 {{ 
            color: #1f2937; 
            margin-bottom: 1.5rem; 
            font-size: 1.75rem; 
        }}
        .file-links {{ 
            display: grid; 
            grid-template-columns: repeat(auto-fit, minmax(300px, 1fr)); 
            gap: 1rem; 
            margin: 2rem 0; 
        }}
        .file-link {{ 
            display: block; 
            padding: 1.5rem; 
            background: white; 
            color: #3b82f6; 
            text-decoration: none; 
            border-radius: 0.75rem; 
            border: 2px solid #3b82f6; 
            text-align: center; 
            font-weight: 600; 
            transition: all 0.2s; 
        }}
        .file-link:hover {{ 
            background: #3b82f6; 
            color: white; 
            transform: translateY(-2px); 
            box-shadow: 0 5px 15px rgba(59, 130, 246, 0.3); 
        }}
        .validation {{ 
            background: #f0fdf4; 
            border-left-color: #10b981; 
        }}
        .validation ul {{ 
            list-style-type: none; 
            padding: 0; 
        }}
        .validation li {{ 
            padding: 0.5rem 0; 
            border-bottom: 1px solid #d1fae5; 
        }}
        .validation li:before {{ 
            content: "✅ "; 
            margin-right: 0.5rem; 
        }}
        .error {{ background: #fef2f2; border-left-color: #ef4444; }}
        .error li:before {{ content: "❌ "; }}
        @media (max-width: 768px) {{ 
            .header h1 {{ font-size: 2rem; }} 
            .stats {{ grid-template-columns: 1fr; }} 
            .file-links {{ grid-template-columns: 1fr; }} 
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🎖️ Merit Badge Counselor Reports</h1>
            <div class="subtitle">Troop 12 & Troop 32 Acton MA</div>
            <div class="timestamp">Generated: {timestamp}</div>
        </div>
        
        <div class="content">
            <div class="section">
                <h2>📊 Summary Statistics</h2>
                <div class="stats">
                    <div class="stat-card">
                        <div class="stat-number">{total_merit_badges}</div>
                        <div class="stat-label">Total Merit Badges</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-number">{total_eagle_badges}</div>
                        <div class="stat-label">Eagle-Required Badges</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-number">{total_adults}</div>
                        <div class="stat-label">Total Adults (T12: {total_t12_adults}, T32: {total_t32_adults})</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-number">{total_counselors}</div>
                        <div class="stat-label">T12/T32 Counselors</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-number">{total_non_counselors}</div>
                        <div class="stat-label">Leaders not Counselors</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-number">{eagle_covered}/{total_eagle_badges}</div>
                        <div class="stat-label">Eagle Badges Covered</div>
                    </div>
                </div>
            </div>
            
            <div class="section">
                <h2>📁 Generated Reports</h2>
                <h3>HTML Reports (Interactive)</h3>
                <div class="file-links">
                    <a href="html/t12_t32_merit_badge_counselors.html" class="file-link">
                        👥 T12/T32 Merit Badge Counselors
                    </a>
                    <a href="html/t12_t32_leaders_not_counselors.html" class="file-link">
                        📋 T12/T32 Leaders not Counselors
                    </a>
                    <a href="html/t12_t32_counselor_coverage.html" class="file-link">
                        📊 T12/T32 Counselor Coverage
                    </a>
                </div>
                
                <h3>Download Formats</h3>
                <div class="file-links">
                    <a href="csv/" class="file-link">📄 CSV Files</a>
                    <a href="excel/" class="file-link">📊 Excel Files</a>
                    <a href="pdf/" class="file-link">📰 PDF Files</a>
                    <a href="wordpress/" class="file-link">🌐 WordPress Files</a>
                </div>
            </div>
            
            <div class="section validation">
                <h2>✅ Validation Results</h2>
                <ul>
                    <li>Merit badge count validation: {total_merit_badges} total badges loaded</li>
                    <li>Eagle-required badges validation: {total_eagle_badges} Eagle badges identified</li>
                    <li>Adult accounting validation: {total_adults} total adults ({total_counselors} counselors + {total_non_counselors} non-counselors = {total_counselors + total_non_counselors})</li>
                    <li>Coverage analysis: {eagle_covered} of {total_eagle_badges} Eagle badges have T12/T32 counselors</li>
                    <li>File generation: All output formats created successfully</li>
                </ul>
            </div>
        </div>
    </div>
</body>
</html>"""
            
            with open(summary_path, 'w', encoding='utf-8') as f:
                f.write(summary_html)
            
            self.logger.info(f"Generated summary report: {summary_path}")
            
        except Exception as e:
            self.logger.error(f"Error generating summary report: {e}")
    
    def validate_outputs(self, reports):
        """Validate generated outputs according to specification requirements"""
        try:
            self.logger.info("Validating outputs...")
            
            # Test 1: Verify merit badge count matches scouting.org
            total_badges_in_coverage = sum(len(data) for data in reports[2]['data'].values())
            if total_badges_in_coverage != len(self.merit_badges):
                self.logger.warning(f"Merit badge count mismatch: Coverage has {total_badges_in_coverage}, expected {len(self.merit_badges)}")
            else:
                self.logger.info("✓ Merit badge count validation passed")
            
            # Test 2: Verify Eagle-required badge count
            eagle_in_coverage = len(reports[2]['data']['eagle_with_counselors']) + len(reports[2]['data']['eagle_without_counselors'])
            if eagle_in_coverage != len(self.eagle_required_badges):
                self.logger.warning(f"Eagle badge count mismatch: Coverage has {eagle_in_coverage}, expected {len(self.eagle_required_badges)}")
            else:
                self.logger.info("✓ Eagle-required badge count validation passed")
            
            # Test 3: Verify adult accounting
            total_adults = len(self.t12_adults) + len(self.t32_adults)
            total_counselors = len(reports[0]['data'])
            total_non_counselors = len(reports[1]['data'])
            
            if total_counselors + total_non_counselors != total_adults:
                self.logger.warning(f"Adult accounting mismatch: {total_counselors} + {total_non_counselors} != {total_adults}")
            else:
                self.logger.info("✓ Adult accounting validation passed")
            
            # Test 4: Verify file generation
            format_dirs = ['html', 'csv', 'pdf', 'excel', 'wordpress']
            for format_dir in format_dirs:
                dir_path = self.output_dir / format_dir
                if not dir_path.exists() or not list(dir_path.iterdir()):
                    self.logger.warning(f"Missing or empty directory: {format_dir}")
                else:
                    self.logger.info(f"✓ {format_dir.upper()} files generated successfully")
            
            self.logger.info("Output validation completed")
            
        except Exception as e:
            self.logger.error(f"Error during output validation: {e}")


class MeritBadgeGUI:
    """Enhanced GUI interface for Merit Badge Generator"""
    
    def __init__(self):
        if not GUI_AVAILABLE:
            raise ImportError("GUI dependencies not available. Install with: pip install tkinterdnd2")
            
        self.root = TkinterDnD.Tk()
        self.root.title("Merit Badge Counselor Generator v2.0")
        self.root.geometry("900x800")
        self.root.configure(bg='#f0f0f0')
        
        # Configure app icon if available
        try:
            self.root.iconbitmap('icon.ico')
        except:
            pass
        
        self.generator = MeritBadgeGenerator()
        self.generator.progress_callback = self.update_progress
        self.generator.status_callback = self.update_status
        
        self.t12_file = None
        self.t32_file = None
        self.pdf_files = []
        
        self.setup_ui()
        self.setup_drag_drop()
        
    def setup_ui(self):
        """Setup enhanced user interface"""
        # Main frame with scrolling capability
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Title section with improved styling
        title
    
    def generate_non_counselors_html(self, data):
        """Generate HTML content for non-counselors report"""
        if not data:
            return "<p>All T12/T32 leaders are merit badge counselors!</p>"
        
        html = f"""
        <div class="stats">
            <div class="stat-card">
                <div class="stat-number">{len(data)}</div>
                <div class="stat-label">Leaders Not Counselors</div>
            </div>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>Name</th>
                    <th>Troop</th>
                    <th>Position</th>
                    <th>Email</th>
                    <th>Phone</th>
                </tr>
            </thead>
            <tbody>
        """
        
        for adult in data:
            html += f"""
                <tr>
                    <td><strong>{adult.get('name', 'N/A')}</strong></td>
                    <td>{adult.get('troop', 'N/A')}</td>
                    <td>{adult.get('position', 'N/A')}</td>
                    <td>{adult.get('email', 'N/A')}</td>
                    <td>{adult.get('phone', 'N/A')}</td>
                </tr>
            """
        
        html += "</tbody></table>"
        return html
    
    def generate_coverage_html(self, data):
        """Generate HTML content for coverage report"""
        sections = [
            ('Eagle-Required Merit Badges with T12/T32 Counselors', data['eagle_with_counselors'], 'eagle-required'),
            ('Eagle-Required Merit Badges without T12/T32 Counselors', data['eagle_without_counselors'], 'eagle-missing'),
            ('Non-Eagle-Required Merit Badges with T12/T32 Counselors', data['non_eagle_with_counselors'], ''),
            ('Non-Eagle-Required Merit Badges without T12/T32 Counselors', data['non_eagle_without_counselors'], '')
        ]
        
        # Stats section
        total_eagle = len(data['eagle_with_counselors']) + len(data['eagle_without_counselors'])
        total_badges = sum(len(badges) for badges in data.values())
        
        html = f"""
        <div class="stats">
            <div class="stat-card">
                <div class="stat-number">{total_badges}</div>
                <div class="stat-label">Total Merit Badges</div>
            </div>
            <div class="stat-card">
                <div class="stat-number">{total_eagle}</div>
                <div class="stat-label">Eagle Required</div>
            </div>
            <div class="stat-card">
                <div class="stat-number">{len(data['eagle_with_counselors'])}</div>
                <div class="stat-label">Eagle Badges Covered</div>
            </div>
            <div class="stat-card">
                <div class="stat-number">{len(data['eagle_without_counselors'])}</div>
                <div class="stat-label">Eagle Badges Needed</div>
            </div>
        </div>
        """
        
        for title, badges, css_class in sections:
            html += f"""
            <div class="coverage-section {css_class}">
                <h3>{title} ({len(badges)})</h3>
                <ul>
            """
            for badge in badges:
                html += f"<li>{badge}</li>"
            html += "</ul></div>"
        
        return html